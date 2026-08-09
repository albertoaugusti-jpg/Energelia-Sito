#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM ENERGELIA — sezione /crm del sito energelia.it
==================================================
Blueprint Flask autonomo. Non tocca il resto di main2.py.

INNESTO (due righe in main2.py, dopo la creazione di `app`):

    from crm import init_crm
    init_crm(app)

Usa lo stesso DATABASE_URL del sito e crea le proprie tabelle, tutte con
prefisso crm_ (crm_utenti, crm_clienti, crm_pratiche, crm_attivita): le
tabelle leads e consultations del sito non vengono toccate.

VARIABILI D'AMBIENTE
    DATABASE_URL     già impostata sul sito.
    CRM_SECRET_KEY   firma dei cookie di sessione. Impostala su Render:
                     senza, viene derivata in modo stabile da DATABASE_URL,
                     che funziona ma è meno solido.
    CRM_ADMIN_EMAIL  primo amministratore (default a.augusti@energelia.it)
    CRM_ADMIN_PASSWORD  password del primo amministratore (default energelia2026)
"""

import os
import io
import csv
import json
import hmac
import hashlib
import secrets
import datetime as dt
from decimal import Decimal, InvalidOperation

import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from flask import (Blueprint, request, session, redirect, url_for,
                   Response, send_file, abort)
from jinja2 import Environment, DictLoader, select_autoescape

from sqlalchemy import (create_engine, Column, Integer, String, Text, Date, DateTime,
                        Numeric, Boolean, ForeignKey, func, or_)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, scoped_session

import openpyxl
import pdfplumber
import pypdf
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------------
# CONFIGURAZIONE
# --------------------------------------------------------------------------

DB_URL = os.environ.get("DATABASE_URL", "sqlite:///crm_energelia.db")
if DB_URL.startswith("postgres://"):          # Render fornisce ancora il vecchio schema
    DB_URL = DB_URL.replace("postgres://", "postgresql://", 1)

ADMIN_EMAIL = os.environ.get("CRM_ADMIN_EMAIL", "a.augusti@energelia.it")
ADMIN_PASSWORD = os.environ.get("CRM_ADMIN_PASSWORD", "energelia2026")

# Google Drive: i clienti caricano documenti tramite un link pubblico, i file
# finiscono nel Drive di Energelia (non del cliente), una cartella per cliente.
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN")
GOOGLE_CARTELLA_MADRE = os.environ.get("GOOGLE_DRIVE_CARTELLA_MADRE")

# Stesso SMTP già usato dal sito per le notifiche lead — nessuna variabile nuova.
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")

# Chiave separata per il repository bandi (diversa da quella dell'Offertatore desktop).
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
ANTHROPIC_MODELLO = "claude-sonnet-5"

engine = create_engine(DB_URL, pool_pre_ping=True, pool_recycle=280,
                       connect_args={"check_same_thread": False} if DB_URL.startswith("sqlite") else {})
# scoped_session: una sessione per thread, chiusa a fine richiesta da init_crm.
SessionLocale = scoped_session(sessionmaker(bind=engine, autoflush=False, expire_on_commit=False))
Base = declarative_base()

# Liste chiuse: identiche a quelle del foglio, per non perdere continuità.
FASI = ["Analisi fattibilità", "Preventivo/Contratto Inviato", "Preventivo/Contratto Firmato",
        "Preparazione documenti", "Presentata", "In graduatoria", "Ammessa",
        "Respinta", "In rendicontazione", "Chiusa"]
FASI_APERTE = [f for f in FASI if f not in ("Respinta", "Chiusa")]      # ancora in lavorazione
FASI_PRE_INVIO = FASI[:4]                                              # domanda non ancora presentata
# Etichette corte per la pipeline: il nome pieno resta quello del dato.
FASI_BREVI = {"Analisi fattibilità": "Analisi", "Preventivo/Contratto Inviato": "Preventivo inviato",
              "Preventivo/Contratto Firmato": "Contratto firmato",
              "Preparazione documenti": "Documenti", "In rendicontazione": "Rendicontazione"}
FASI_VINTE = ["Ammessa", "In rendicontazione", "Chiusa"]
CANALI = ["Sito Web", "Referral", "Chiamata", "Instagram", "Fiera/Evento", "Passaparola", "Scrapping", "Altro"]
TIPOLOGIE = ["Fondo perduto", "Finanziamento agevolato", "Credito d'imposta", "Voucher", "Misto"]
PRIORITA = ["Alta", "Media", "Bassa"]
DIMENSIONI = ["Micro", "Piccola", "Media", "Grande"]
TIPI_ATTIVITA = ["Chiamata", "Email", "Riunione", "Documenti", "Nota"]
STATI_LEAD = ["nuovo", "contattato", "scartato", "convertito"]
RUOLI = ["admin", "consulente"]

# --------------------------------------------------------------------------
# MODELLI
# --------------------------------------------------------------------------

class Utente(Base):
    __tablename__ = "crm_utenti"
    id = Column(Integer, primary_key=True)
    nome = Column(String(120), nullable=False)
    email = Column(String(200), unique=True, nullable=False, index=True)
    password_hash = Column(String(255), nullable=False)
    ruolo = Column(String(20), default="consulente")
    attivo = Column(Boolean, default=True)
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    @property
    def is_admin(self):
        return self.ruolo == "admin"


class ContoBancario(Base):
    """I conti di Energelia su cui i clienti versano il corrispettivo.
    Gestiti una volta sola in Impostazioni, poi scelti con un menu a tendina su ogni pratica."""
    __tablename__ = "crm_conti_bancari"
    id = Column(Integer, primary_key=True)
    nome = Column(String(120), nullable=False)   # es. "Energelia — Ordinario"
    iban = Column(String(34))
    banca = Column(String(120))
    attivo = Column(Boolean, default=True)


class Cliente(Base):
    __tablename__ = "crm_clienti"
    id = Column(Integer, primary_key=True)
    codice = Column(String(20), unique=True, index=True)
    ragione_sociale = Column(String(255), nullable=False)
    piva = Column(String(40))
    ateco = Column(String(40))
    dimensione = Column(String(20))
    citta = Column(String(120))
    provincia = Column(String(10))
    regione = Column(String(60))
    referente = Column(String(160))
    ruolo_referente = Column(String(120))
    telefono = Column(String(60))
    email = Column(String(200))
    pec = Column(String(200))
    iban = Column(String(34))                  # IBAN del cliente: dove riceve l'accredito del bando
    codice_fiscale = Column(String(16))
    codice_sdi = Column(String(10))             # fatturazione elettronica
    intestatario_conto = Column(String(255))    # se diverso dalla ragione sociale
    email_fatturazione = Column(String(200))
    titolari_effettivi = Column(Text)   # nomi, non la composizione societaria completa
    token_caricamento = Column(String(64), unique=True, index=True)  # link pubblico di upload
    google_cartella_id = Column(String(120))                         # cartella Drive di questo cliente
    consulente_id = Column(Integer, ForeignKey("crm_utenti.id"))
    canale = Column(String(40))
    data_primo_contatto = Column(Date)
    prossima_azione = Column(String(255))
    note = Column(Text)
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    consulente = relationship("Utente")
    pratiche = relationship("Pratica", back_populates="cliente",
                            cascade="all, delete-orphan", order_by="Pratica.codice")
    attivita = relationship("Attivita", back_populates="cliente",
                            cascade="all, delete-orphan",
                            order_by="Attivita.data.desc(), Attivita.id.desc()")

    @property
    def ultimo_contatto(self):
        return self.attivita[0].data if self.attivita else None


class Pratica(Base):
    __tablename__ = "crm_pratiche"
    id = Column(Integer, primary_key=True)
    codice = Column(String(20), unique=True, index=True)
    cliente_id = Column(Integer, ForeignKey("crm_clienti.id"), nullable=False)
    bando_id = Column(Integer, ForeignKey("crm_bandi.id"))   # da quale bando del repository nasce, se c'è
    nome_bando = Column(String(255), nullable=False)
    ente = Column(String(160))
    tipologia = Column(String(60))
    perc_contributo = Column(Numeric(6, 2))          # in punti percentuali: 75.00
    importo_max = Column(String(120))                # testo libero: "30k linea A / 40k linea B"
    data_apertura = Column(Date)
    data_scadenza = Column(Date)
    fase = Column(String(60), default="Analisi fattibilità")
    data_presentazione = Column(Date)
    data_esito = Column(Date)
    importo_richiesto = Column(Numeric(14, 2))
    importo_concesso = Column(Numeric(14, 2))
    scadenza_rendicontazione = Column(Date)
    documenti_mancanti = Column(Text)
    prossimo_step = Column(String(255))
    priorita = Column(String(20), default="Media")
    corrispettivo = Column(Numeric(12, 2))
    success_fee_perc = Column(Numeric(6, 2))
    fatturato = Column(Numeric(12, 2))
    incassato = Column(Numeric(12, 2))
    conto_incasso_id = Column(Integer, ForeignKey("crm_conti_bancari.id"))
    note = Column(Text)
    token_caricamento = Column(String(64), unique=True, index=True)   # link pubblico per QUESTA pratica
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    cliente = relationship("Cliente", back_populates="pratiche")
    attivita = relationship("Attivita", back_populates="pratica")
    conto_incasso = relationship("ContoBancario")
    bando = relationship("Bando")
    voci_richiesta = relationship("VoceRichiesta", back_populates="pratica",
                                  order_by="VoceRichiesta.ordine", cascade="all, delete-orphan")

    @property
    def success_fee_maturata(self):
        """Success fee sull'importo concesso. Zero finché non c'è un concesso."""
        if self.importo_concesso and self.success_fee_perc:
            return self.importo_concesso * self.success_fee_perc / Decimal(100)
        return Decimal(0)

    @property
    def compenso_totale(self):
        return (self.corrispettivo or Decimal(0)) + self.success_fee_maturata

    @property
    def giorni_a_scadenza(self):
        """Giorni alla scadenza di presentazione. None se la domanda è già partita."""
        if not self.data_scadenza or self.fase not in FASI_PRE_INVIO:
            return None
        return (self.data_scadenza - dt.date.today()).days


class Attivita(Base):
    __tablename__ = "crm_attivita"
    id = Column(Integer, primary_key=True)
    cliente_id = Column(Integer, ForeignKey("crm_clienti.id"), nullable=False)
    pratica_id = Column(Integer, ForeignKey("crm_pratiche.id"))
    data = Column(Date, default=dt.date.today, nullable=False)
    tipo = Column(String(40), default="Nota")
    testo = Column(Text, nullable=False)
    utente_id = Column(Integer, ForeignKey("crm_utenti.id"))
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    cliente = relationship("Cliente", back_populates="attivita")
    pratica = relationship("Pratica", back_populates="attivita")
    utente = relationship("Utente")


class Lead(Base):
    """Nominativo generico dagli scraper — leggero apposta, pensato per volumi
    grandi (decine di migliaia). Diventa Cliente con una conversione esplicita,
    non è collegato a nient'altro finché qualcuno non lo promuove."""
    __tablename__ = "crm_lead"
    id = Column(Integer, primary_key=True)
    nome = Column(String(255), nullable=False, index=True)
    tipo = Column(String(120))               # categoria Maps, es. "Ristorante italiano"
    query_ricerca = Column(String(200))      # keyword usata dallo scraper
    indirizzo = Column(Text)      # senza limite: gli indirizzi reali variano parecchio
    comune = Column(String(120))
    provincia = Column(String(10), index=True)
    cap = Column(String(10))
    telefono = Column(String(60))
    cellulare = Column(String(60))
    sito = Column(Text)           # senza limite: alcuni URL con parametri sono lunghissimi
    email = Column(String(200), index=True)
    altra_email = Column(String(200))
    pec = Column(String(200))
    fonte = Column(String(200))              # etichetta del giro di scraping, es. "Valle d'Aosta - Ristoranti"
    stato = Column(String(20), default="nuovo")   # nuovo / contattato / scartato / convertito
    cliente_id = Column(Integer, ForeignKey("crm_clienti.id"))
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    cliente = relationship("Cliente")

    @property
    def contatto_migliore(self):
        return self.cellulare or self.telefono or self.email or "—"


class Documento(Base):
    """Un file caricato dal cliente (o da voi) tramite il link pubblico.
    Il file vero sta su Google Drive: qui teniamo solo il riferimento e lo
    stato di smistamento — non ancora assegnato a una pratica, o già assegnato."""
    __tablename__ = "crm_documenti"
    id = Column(Integer, primary_key=True)
    cliente_id = Column(Integer, ForeignKey("crm_clienti.id"), nullable=False)
    pratica_id = Column(Integer, ForeignKey("crm_pratiche.id"))
    nome_file = Column(String(300), nullable=False)
    google_file_id = Column(String(120))
    link_drive = Column(String(500))
    dimensione_byte = Column(Integer)
    caricato_da = Column(String(20), default="cliente")   # cliente / staff
    stato = Column(String(20), default="da_smistare")      # da_smistare / assegnato
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    cliente = relationship("Cliente")
    pratica = relationship("Pratica")


class VoceRichiesta(Base):
    """Una voce specifica che chiedete al cliente per QUESTA pratica —
    testo libero (es. una spiegazione) oppure uno slot di caricamento con
    nome (es. 'Preventivo fornitore'), distinta dal cestello generico."""
    __tablename__ = "crm_voci_richiesta"
    id = Column(Integer, primary_key=True)
    pratica_id = Column(Integer, ForeignKey("crm_pratiche.id"), nullable=False)
    etichetta = Column(String(200), nullable=False)
    tipo_risposta = Column(String(10), default="file")   # file / testo
    valore_testo = Column(Text)
    documento_id = Column(Integer, ForeignKey("crm_documenti.id"))
    compilata = Column(Boolean, default=False)
    ordine = Column(Integer, default=0)
    creato_il = Column(DateTime, default=dt.datetime.utcnow)

    pratica = relationship("Pratica", back_populates="voci_richiesta")
    documento = relationship("Documento")


class Bando(Base):
    """Repository dei bandi: si crea a mano o caricando la scheda PDF (che
    un'IA legge e smonta nei campi). Da qui si genera una pratica già
    precompilata, e — a richiesta — una guida di compilazione scaricabile
    e un approfondimento più lungo della scheda base."""
    __tablename__ = "crm_bandi"
    id = Column(Integer, primary_key=True)
    nome = Column(String(255), nullable=False, index=True)
    ente = Column(String(200))
    tipologia = Column(String(60))
    dotazione = Column(Numeric(14, 2))
    perc_contributo = Column(Numeric(6, 2))
    contributo_max = Column(Numeric(14, 2))
    importo_max_testo = Column(String(200))   # es. "30k linea A / 40k linea B"
    data_apertura = Column(Date)
    data_scadenza = Column(Date)
    chi_puo_partecipare = Column(Text)
    cosa_finanziabile = Column(Text)
    spese_non_ammissibili = Column(Text)
    criteri = Column(Text)
    fasi_tempi = Column(Text)
    come_presentare = Column(Text)
    perche_interessante = Column(Text)
    criticita = Column(Text)
    testo_originale = Column(Text)            # testo grezzo della scheda caricata, per rileggerlo
    guida_compilazione = Column(Text)         # generata dall'IA, nullable finché non richiesta
    approfondimento = Column(Text)            # generato dall'IA, nullable finché non richiesto
    creato_il = Column(DateTime, default=dt.datetime.utcnow)


# --------------------------------------------------------------------------
# UTILITÀ
# --------------------------------------------------------------------------

def hash_pw(password: str, salt: str = None) -> str:
    salt = salt or secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
    return f"pbkdf2${salt}${dk.hex()}"


def verifica_pw(password: str, stored: str) -> bool:
    try:
        _, salt, _ = stored.split("$")
    except ValueError:
        return False
    return hmac.compare_digest(hash_pw(password, salt), stored)


def drive_configurato():
    return bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and GOOGLE_REFRESH_TOKEN and GOOGLE_CARTELLA_MADRE)


def smtp_configurato():
    return bool(SMTP_USER and SMTP_PASS)


def invia_email(destinatario, oggetto, corpo, rispondi_a=None, nome_mittente="Energelia"):
    """Manda dalla casella del sito (SMTP_USER), ma con Reply-To sull'utente che
    ha cliccato 'Invia' — così una risposta del cliente arriva a lui, non alla
    casella condivisa. Ritorna (True, '') o (False, messaggio d'errore)."""
    if not smtp_configurato():
        return False, "L'invio email non è configurato su questo sito."
    try:
        msg = MIMEMultipart()
        msg["From"] = f"{nome_mittente} <{SMTP_USER}>"
        msg["To"] = destinatario
        msg["Subject"] = oggetto
        if rispondi_a:
            msg["Reply-To"] = rispondi_a
        msg.attach(MIMEText(corpo, "plain", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
        return True, ""
    except Exception as errore:
        return False, str(errore)


def anthropic_configurato():
    return bool(ANTHROPIC_API_KEY)


def _anthropic_chiama(system, messaggio, max_token=2000, con_ricerca_web=False):
    """Una chiamata semplice all'API Messages. Solleva eccezione se qualcosa
    va storto — chi la usa decide come mostrarlo (avvisa/log).
    Con con_ricerca_web=True l'IA può cercare sul web da sola (strumento
    lato server di Anthropic: la ricerca vera e propria non passa da qui,
    la fa Anthropic e ci restituisce solo il testo finale)."""
    corpo = {
        "model": ANTHROPIC_MODELLO,
        "max_tokens": max_token,
        "system": system,
        "messages": [{"role": "user", "content": messaggio}],
    }
    if con_ricerca_web:
        corpo["tools"] = [{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}]
    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json=corpo,
        timeout=120,
    )
    r.raise_for_status()
    blocchi = r.json().get("content", [])
    return "".join(b.get("text", "") for b in blocchi if b.get("type") == "text")


def _estrai_testo_pdf_pdfplumber(contenuto_bytes):
    testo = []
    pagine_saltate = 0
    with pdfplumber.open(io.BytesIO(contenuto_bytes)) as pdf:
        for numero, pagina in enumerate(pdf.pages, start=1):
            try:
                testo.append(pagina.extract_text() or "")
            except Exception as errore:
                pagine_saltate += 1
                print(f"[crm] pdfplumber: pagina {numero} illeggibile, salto: {errore}")
    return "\n".join(testo).strip(), pagine_saltate


def _estrai_testo_pdf_pypdf(contenuto_bytes):
    lettore = pypdf.PdfReader(io.BytesIO(contenuto_bytes))
    testo = []
    for numero, pagina in enumerate(lettore.pages, start=1):
        try:
            testo.append(pagina.extract_text() or "")
        except Exception as errore:
            print(f"[crm] pypdf: pagina {numero} illeggibile, salto: {errore}")
    return "\n".join(testo).strip()


def _estrai_testo_pdf(contenuto_bytes):
    """Alcuni PDF (spesso esportazioni da sistemi camerali o scansioni) hanno
    strutture interne che fanno inciampare pdfplumber (bug noto di pdfminer,
    tipicamente 'unhashable type: dict'). pdfplumber e pypdf sono due motori
    indipendenti: se il primo fallisce del tutto, provo il secondo prima di
    arrendermi — sono scritti da persone diverse, raramente si rompono sullo
    stesso file per lo stesso motivo."""
    try:
        risultato, pagine_saltate = _estrai_testo_pdf_pdfplumber(contenuto_bytes)
    except Exception as errore:
        print(f"[crm] pdfplumber non apre il file, provo pypdf: {errore}")
        risultato, pagine_saltate = "", None

    if not risultato:
        try:
            risultato = _estrai_testo_pdf_pypdf(contenuto_bytes)
        except Exception as errore:
            print(f"[crm] anche pypdf fallisce: {errore}")

    if not risultato.strip():
        raise ValueError(
            "Non riesco a leggere il testo di questo PDF con nessuno dei due motori "
            "disponibili: potrebbe essere una scansione senza testo selezionabile."
        )
    return risultato.strip()


SYSTEM_ESTRAI_BANDO = """Estrai i dati da una scheda di bando italiana (finanza agevolata).

Rispondi ESCLUSIVAMENTE con un oggetto JSON, senza testo prima o dopo, con queste chiavi:
nome, ente, tipologia, dotazione, perc_contributo, contributo_max, importo_max_testo,
data_apertura, data_scadenza, chi_puo_partecipare, cosa_finanziabile, spese_non_ammissibili,
criteri, fasi_tempi, come_presentare, perche_interessante, criticita

Regole:
- Se un dato non è presente o non sei sicuro, metti stringa vuota "". Non inventare.
- "tipologia" è una di queste: Fondo perduto, Finanziamento agevolato, Credito d'imposta, Voucher, Misto.
- "dotazione", "perc_contributo", "contributo_max": solo numeri (punto decimale, niente simboli),
  o stringa vuota se non ricavabile con certezza. perc_contributo in punti percentuali (75, non 0.75).
- "importo_max_testo": usalo quando il tetto non è un numero secco (es. "30k linea A / 40k linea B").
- Le date in formato YYYY-MM-DD.
- I campi discorsivi (chi_puo_partecipare, cosa_finanziabile, ecc.) sono sintesi in elenco puntato
  testuale (righe con "- "), non frasi uniche.
"""


def estrai_bando_da_testo(testo_scheda):
    if not (testo_scheda or "").strip():
        raise ValueError("Il testo estratto dal PDF è vuoto: file scansionato o illeggibile.")
    grezzo = _anthropic_chiama(SYSTEM_ESTRAI_BANDO, testo_scheda[:20000], max_token=2500)
    grezzo = grezzo.strip()
    if grezzo.startswith("```"):
        grezzo = grezzo.split("\n", 1)[1].rsplit("```", 1)[0]
    return json.loads(grezzo)


SYSTEM_GUIDA_COMPILAZIONE = """Sei un consulente di finanza agevolata italiano di Energelia S.r.l.
Scrivi una guida pratica alla compilazione della domanda per il bando che ti viene descritto,
pensata per un piccolo imprenditore che deve presentarla da solo o con il vostro supporto.

Struttura in sezioni con intestazioni chiare:
1. Prima di iniziare — cosa procurarsi (documenti, credenziali SPID/CIE/CNS, dati aziendali)
2. Passo per passo — come si compila, in ordine, con attenzione ai punti dove si sbaglia più spesso
3. Errori da evitare — nello specifico di QUESTO bando, non generici
4. Dopo l'invio — cosa aspettarsi (istruttoria, tempi, rendicontazione)

Tono diretto e concreto, frasi brevi. Non ripetere la scheda del bando, dai per scontato che il
lettore l'abbia già letta: questa è la guida operativa, non un riassunto del bando."""


def genera_guida_compilazione(bando):
    descrizione = (
        f"Bando: {bando.nome}\nEnte: {bando.ente or '—'}\nTipologia: {bando.tipologia or '—'}\n"
        f"Come si presenta: {bando.come_presentare or '—'}\nCriteri: {bando.criteri or '—'}\n"
        f"Fasi e tempi: {bando.fasi_tempi or '—'}\nCriticità note: {bando.criticita or '—'}\n"
        f"Testo originale della scheda:\n{(bando.testo_originale or '')[:15000]}"
    )
    return _anthropic_chiama(SYSTEM_GUIDA_COMPILAZIONE, descrizione, max_token=3000)


SYSTEM_APPROFONDIMENTO = """Sei un consulente di finanza agevolata italiano di Energelia S.r.l.
Scrivi un approfondimento più esteso e tecnico di questo bando rispetto alla scheda sintetica che
già esiste — per un cliente che sta valutando se e come farne richiesta insieme a voi.

Vai oltre i punti già coperti nella scheda base: casi limite, interazione con altri incentivi
(cumulabilità), interpretazioni non ovvie dei requisiti, cosa chiedere prima di procedere.
Se un punto non è chiaro dal testo della scheda, dillo esplicitamente invece di inventare.
Tono diretto, niente fronzoli, frasi brevi."""


def genera_approfondimento(bando):
    descrizione = f"Bando: {bando.nome}\n\nTesto della scheda:\n{(bando.testo_originale or '')[:15000]}"
    return _anthropic_chiama(SYSTEM_APPROFONDIMENTO, descrizione, max_token=3000)


SYSTEM_ESTRAI_VISURA = """Estrai dati anagrafici da una visura camerale italiana.

Rispondi ESCLUSIVAMENTE con un oggetto JSON, senza testo prima o dopo, con queste chiavi:
ragione_sociale, piva, codice_fiscale, ateco, citta, provincia, regione, pec,
referente, ruolo_referente, titolari_effettivi

Regole:
- Se un dato non è presente o non sei sicuro, metti stringa vuota "". Non inventare.
- "pec": l'indirizzo di posta elettronica certificata, se presente nella visura.
- "referente" è il nome del legale rappresentante (chi ha i poteri di firma), non un socio qualunque.
- "ruolo_referente" per il legale rappresentante è di norma "Legale rappresentante" o la carica esatta
  se diversa (es. "Amministratore Unico", "Presidente CdA").
- "titolari_effettivi": SOLO i nomi delle persone fisiche indicate come titolari effettivi, uno per
  riga con "- " davanti. NON includere l'elenco completo della compagine societaria, le quote di
  partecipazione o altri soci che non sono titolari effettivi: quella parte va ignorata del tutto.
- "citta"/"provincia"/"regione" si riferiscono alla sede legale.
"""


def estrai_visura_da_testo(testo_visura):
    if not (testo_visura or "").strip():
        raise ValueError("Il testo estratto dal PDF è vuoto: file scansionato o illeggibile.")
    grezzo = _anthropic_chiama(SYSTEM_ESTRAI_VISURA, testo_visura[:20000], max_token=1500).strip()
    if grezzo.startswith("```"):
        grezzo = grezzo.split("\n", 1)[1].rsplit("```", 1)[0]
    return json.loads(grezzo)


SYSTEM_CERCA_CLIENTE = """Trova informazioni pubbliche di contatto su un'azienda italiana usando la
ricerca web (Google Maps, il sito dell'azienda, elenchi pubblici tipo Registro Imprese/PagineGialle).

Rispondi ESCLUSIVAMENTE con un oggetto JSON, senza testo prima o dopo, con queste chiavi:
sito, telefono, email, pec, indirizzo, citta, provincia

Regole:
- Se non trovi un dato con ragionevole certezza, metti stringa vuota "". Non inventare MAI un numero,
  un indirizzo o un'email che non hai trovato in una fonte.
- Verifica che il risultato corrisponda davvero all'azienda indicata (stessa città/provincia se data,
  non un'azienda omonima in un'altra zona) prima di riportare un dato.
- "pec" solo se la trovi esplicitamente indicata come PEC (posta elettronica certificata), non un'email
  generica scambiata per PEC.
"""


def cerca_cliente_online(cliente):
    pezzi = [f"Ragione sociale: {cliente.ragione_sociale}"]
    if cliente.citta: pezzi.append(f"Città: {cliente.citta}")
    if cliente.provincia: pezzi.append(f"Provincia: {cliente.provincia}")
    if cliente.piva: pezzi.append(f"Partita IVA: {cliente.piva}")
    messaggio = "\n".join(pezzi)

    grezzo = _anthropic_chiama(SYSTEM_CERCA_CLIENTE, messaggio, max_token=1500, con_ricerca_web=True).strip()
    if grezzo.startswith("```"):
        grezzo = grezzo.split("\n", 1)[1].rsplit("```", 1)[0]
    # con la ricerca web l'IA a volte aggiunge una frase prima/dopo il JSON nonostante l'istruzione:
    # prendo solo la parte fra la prima { e l'ultima }
    inizio, fine = grezzo.find("{"), grezzo.rfind("}")
    if inizio == -1 or fine == -1:
        raise ValueError("La ricerca non ha restituito un risultato leggibile.")
    return json.loads(grezzo[inizio:fine + 1])


def _drive_token_accesso():
    r = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id": GOOGLE_CLIENT_ID, "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN, "grant_type": "refresh_token",
    }, timeout=15)
    r.raise_for_status()
    return r.json()["access_token"]


def _drive_cartella_cliente(cliente):
    """Ritorna l'ID della cartella Drive del cliente, creandola al primo utilizzo."""
    if cliente.google_cartella_id:
        return cliente.google_cartella_id
    token = _drive_token_accesso()
    nome = f"{cliente.codice} — {cliente.ragione_sociale}"
    r = requests.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": nome, "mimeType": "application/vnd.google-apps.folder",
              "parents": [GOOGLE_CARTELLA_MADRE]},
        timeout=15)
    r.raise_for_status()
    cartella_id = r.json()["id"]
    cliente.google_cartella_id = cartella_id
    SessionLocale.commit()
    return cartella_id


def _drive_carica_file(cartella_id, nome_file, contenuto, mimetype):
    token = _drive_token_accesso()
    metadata = {"name": nome_file, "parents": [cartella_id]}
    files = {
        "metadata": ("metadata", json.dumps(metadata), "application/json"),
        "file": (nome_file, contenuto, mimetype or "application/octet-stream"),
    }
    r = requests.post(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=multipart&fields=id,webViewLink,size",
        headers={"Authorization": f"Bearer {token}"}, files=files, timeout=60)
    r.raise_for_status()
    return r.json()


def prossimo_codice(db, modello, prefisso):
    ultimo = db.query(func.max(modello.codice)).scalar()
    n = 1
    if ultimo and "-" in ultimo:
        try:
            n = int(ultimo.split("-")[1]) + 1
        except ValueError:
            n = db.query(func.count(modello.id)).scalar() + 1
    return f"{prefisso}-{n:04d}"


def d(valore):
    """Converte una stringa in data. Accetta 2026-03-31 e 31/03/2026."""
    if not valore:
        return None
    valore = str(valore).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(valore, fmt).date()
        except ValueError:
            continue
    return None


def n(valore):
    """Converte in numero. Accetta 30.000,50 e 30000.50. Stringa vuota -> None."""
    if valore is None:
        return None
    if isinstance(valore, (int, float, Decimal)):
        return Decimal(str(valore))
    testo = str(valore).strip().replace("€", "").replace("%", "").replace(" ", "").replace("\u00a0", "")
    if not testo:
        return None
    if "," in testo and "." in testo:
        # 30.000,50 -> il punto separa le migliaia
        testo = testo.replace(".", "").replace(",", ".")
    elif "," in testo:
        testo = testo.replace(",", ".")
    elif "." in testo:
        # Notazione italiana: 8.000 sono ottomila, non otto. Il punto vale come
        # separatore di migliaia solo se ogni gruppo dopo il primo ha 3 cifre.
        gruppi = testo.lstrip("-+").split(".")
        if len(gruppi) > 1 and all(len(g) == 3 and g.isdigit() for g in gruppi[1:]) \
                and gruppi[0].isdigit() and 1 <= len(gruppi[0]) <= 3:
            testo = testo.replace(".", "")
    try:
        return Decimal(testo)
    except InvalidOperation:
        return None


def s(valore):
    """Normalizza una stringa di form: vuota -> None."""
    if valore is None:
        return None
    valore = str(valore).strip()
    return valore or None


def euro(valore):
    if valore in (None, ""):
        return "—"
    return f"{Decimal(valore):,.0f}".replace(",", ".") + " €"


def data_it(valore):
    return valore.strftime("%d/%m/%Y") if valore else "—"


# --------------------------------------------------------------------------
# TEMPLATE
# --------------------------------------------------------------------------

BASE = """<!doctype html>
<html lang="it"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ titolo }} · CRM Energelia</title>
<style>
:root{
  --navy:#1f4e78; --navy-scuro:#143453; --ambra:#e07a2f; --carta:#f6f7f9;
  --bordo:#dde2e8; --testo:#1b2733; --tenue:#6b7b8c; --verde:#1f7a4d; --rosso:#c0392b;
}
*{box-sizing:border-box}
body{margin:0;background:var(--carta);color:var(--testo);
  font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif}
a{color:var(--navy);text-decoration:none} a:hover{text-decoration:underline}
.etichetta{font-size:10px;letter-spacing:.10em;text-transform:uppercase;color:var(--tenue);font-weight:700}

header{background:var(--navy);color:#fff;padding:0 20px;display:flex;align-items:center;
  gap:26px;flex-wrap:wrap;box-shadow:0 1px 0 rgba(0,0,0,.12)}
header .marchio{font-weight:800;letter-spacing:.14em;font-size:13px;padding:14px 0;white-space:nowrap}
header .marchio span{color:var(--ambra)}
header nav{display:flex;gap:4px;flex:1;flex-wrap:wrap}
header nav a{color:#c9d8e6;padding:14px 12px;font-weight:600;font-size:13px;border-bottom:3px solid transparent}
header nav a:hover{color:#fff;text-decoration:none}
header nav a.attivo{color:#fff;border-bottom-color:var(--ambra)}
header .utente{font-size:12px;color:#a9c0d4;padding:14px 0}
header .utente a{color:#fff}

main{max-width:1400px;margin:0 auto;padding:22px 20px 60px}
h1{font-size:21px;margin:0 0 2px;font-weight:700;letter-spacing:-.01em}
h2{font-size:14px;margin:26px 0 10px;font-weight:700}
.sottotitolo{color:var(--tenue);font-size:13px;margin:0 0 18px}
.testa{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap}

.riquadro{background:#fff;border:1px solid var(--bordo);border-radius:6px}
.riquadro .corpo{padding:16px}
.griglia{display:grid;gap:14px}
.g4{grid-template-columns:repeat(4,1fr)} .g3{grid-template-columns:repeat(3,1fr)}
.g2{grid-template-columns:repeat(2,1fr)}
@media(max-width:900px){.g4,.g3,.g2{grid-template-columns:1fr 1fr}}
@media(max-width:560px){.g4,.g3,.g2{grid-template-columns:1fr}}

.kpi{background:#fff;border:1px solid var(--bordo);border-radius:6px;padding:14px 16px}
.kpi .valore{font-size:27px;font-weight:750;letter-spacing:-.02em;margin-top:6px;
  font-variant-numeric:tabular-nums;line-height:1.1}
.kpi .nota{font-size:11px;color:var(--tenue);margin-top:4px}

/* Pipeline: le fasi sono una sequenza vera, quindi la barra le mostra in ordine. */
.pipeline{display:flex;border:1px solid var(--bordo);border-radius:6px;overflow:hidden;background:#fff}
.pipeline a{flex:1;padding:12px 10px;border-right:1px solid var(--bordo);min-width:0;
  display:block;color:inherit}
.pipeline a:last-child{border-right:0}
.pipeline a:hover{background:#f0f4f8;text-decoration:none}
.pipeline .num{font-size:22px;font-weight:750;font-variant-numeric:tabular-nums;line-height:1}
.pipeline .nome{font-size:10px;color:var(--tenue);margin-top:5px;line-height:1.25;
  text-transform:uppercase;letter-spacing:.03em;font-weight:600;
  overflow-wrap:break-word}
.pipeline .vuota .num{color:#c3ccd6}
.pipeline .barra{height:3px;background:var(--ambra);margin-top:9px;border-radius:2px}
@media(max-width:900px){.pipeline{flex-wrap:wrap}.pipeline a{flex:1 1 33%}}

table{width:100%;border-collapse:collapse;background:#fff;font-size:13px}
th{text-align:left;padding:9px 12px;background:#eef1f5;border-bottom:1px solid var(--bordo);
  font-size:10px;letter-spacing:.08em;text-transform:uppercase;color:#54677a;font-weight:700;white-space:nowrap}
td{padding:10px 12px;border-bottom:1px solid #eef1f4;vertical-align:top}
tbody tr:hover{background:#fafbfc}
td.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.tabella{border:1px solid var(--bordo);border-radius:6px;overflow:hidden}
.scorri{overflow-x:auto}

.pill{display:inline-block;padding:2px 8px;border-radius:11px;font-size:11px;font-weight:650;
  border:1px solid transparent;white-space:nowrap}
.p-alta{background:#fdecea;color:#a12d20;border-color:#f5c6c0}
.p-media{background:#fdf3e3;color:#8a5518;border-color:#f2ddb9}
.p-bassa{background:#eef1f5;color:#5a6b7c;border-color:#dde2e8}
.fase{background:#e9f0f7;color:var(--navy-scuro);border-color:#cfdeeb}
.fase-ok{background:#e6f4ec;color:var(--verde);border-color:#c3e4d2}
.fase-ko{background:#f2f4f6;color:#7b8a99;border-color:#e0e5ea}
.scad-rossa{color:var(--rosso);font-weight:700}
.scad-ambra{color:#a8621b;font-weight:650}

.btn{display:inline-block;background:var(--navy);color:#fff;border:1px solid var(--navy);
  padding:8px 15px;border-radius:5px;font-size:13px;font-weight:650;cursor:pointer}
.btn:hover{background:var(--navy-scuro);text-decoration:none;color:#fff}
.btn.chiaro{background:#fff;color:var(--navy)} .btn.chiaro:hover{background:#eef2f7;color:var(--navy)}
.btn.ambra{background:var(--ambra);border-color:var(--ambra)} .btn.ambra:hover{background:#c8681f}

form.filtri{display:flex;gap:8px;flex-wrap:wrap;margin:0 0 14px;align-items:center}
input,select,textarea{font:inherit;padding:7px 9px;border:1px solid var(--bordo);border-radius:5px;
  background:#fff;color:var(--testo);max-width:100%}
input:focus,select:focus,textarea:focus{outline:2px solid var(--navy);outline-offset:-1px;border-color:var(--navy)}
textarea{width:100%;min-height:72px;resize:vertical}
label{display:block;margin-bottom:14px}
label .etichetta{display:block;margin-bottom:4px}
label input,label select{width:100%}
.campo-pw{position:relative}
.campo-pw input{padding-right:34px;width:100%;box-sizing:border-box}
.campo-pw .occhio{position:absolute;right:4px;top:50%;transform:translateY(-50%);background:none;
  border:0;cursor:pointer;font-size:15px;line-height:1;padding:6px;color:var(--tenue)}
.campo-pw .occhio:hover{color:var(--navy)}
fieldset{border:1px solid var(--bordo);border-radius:6px;padding:16px 16px 2px;margin:0 0 16px;background:#fff}
legend{font-size:10px;letter-spacing:.10em;text-transform:uppercase;color:var(--tenue);
  font-weight:700;padding:0 7px}

.avviso{padding:11px 14px;border-radius:5px;margin-bottom:16px;font-size:13px;border:1px solid}
.ok{background:#e8f5ee;border-color:#bfe0cd;color:#1c6340}
.ko{background:#fdecea;border-color:#f2c4bd;color:#a3271a}
.vuoto{padding:40px 20px;text-align:center;color:var(--tenue);background:#fff;
  border:1px dashed var(--bordo);border-radius:6px}
.dettaglio{display:grid;grid-template-columns:190px 1fr;gap:1px;background:var(--bordo);
  border:1px solid var(--bordo);border-radius:6px;overflow:hidden}
.dettaglio dt{background:#f4f6f9;padding:9px 12px;font-size:11px;font-weight:700;color:#54677a;
  text-transform:uppercase;letter-spacing:.05em}
.dettaglio dd{background:#fff;padding:9px 12px;margin:0}
@media(max-width:560px){.dettaglio{grid-template-columns:1fr}.dettaglio dt{padding-bottom:2px}}
.diario{list-style:none;padding:0;margin:0}
.diario li{background:#fff;border:1px solid var(--bordo);border-radius:6px;padding:11px 14px;margin-bottom:8px}
.diario .capo{display:flex;gap:10px;align-items:center;font-size:11px;color:var(--tenue);margin-bottom:5px;flex-wrap:wrap}
.diario .tipo{font-weight:700;color:var(--navy);text-transform:uppercase;letter-spacing:.05em}
@media(prefers-reduced-motion:no-preference){.kpi,.riquadro{transition:box-shadow .15s}}
</style></head><body>
{% if utente %}
<header>
  <div class="marchio">ENERGELIA<span>·</span>CRM</div>
  <nav>
    <a href="/crm/" class="{{ 'attivo' if pagina=='dashboard' }}">Dashboard</a>
    <a href="/crm/clienti" class="{{ 'attivo' if pagina=='clienti' }}">Clienti</a>
    <a href="/crm/pratiche" class="{{ 'attivo' if pagina=='pratiche' }}">Pratiche</a>
    <a href="/crm/attivita" class="{{ 'attivo' if pagina=='attivita' }}">Attività</a>
    <a href="/crm/lead" class="{{ 'attivo' if pagina=='lead' }}">Lead</a>
    <a href="/crm/documenti" class="{{ 'attivo' if pagina=='documenti' }}">Documenti</a>
    <a href="/crm/bandi" class="{{ 'attivo' if pagina=='bandi' }}">Bandi</a>
    {% if utente.is_admin %}<a href="/crm/impostazioni" class="{{ 'attivo' if pagina=='impostazioni' }}">Impostazioni</a>{% endif %}
  </nav>
  <div class="utente">{{ utente.nome }} · <a href="/crm/esci">Esci</a></div>
</header>
{% endif %}
<main>
{% for tipo, testo in messaggi %}<div class="avviso {{ tipo }}">{{ testo }}</div>{% endfor %}
{% block contenuto %}{% endblock %}
</main>
<script>
function mostraPw(id){var c=document.getElementById(id);c.type=(c.type==='password')?'text':'password';}
</script>
</body></html>"""

T_LOGIN = """<!doctype html><html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Accedi · CRM Energelia</title>
<style>
body{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;
 background:#1f4e78;font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;padding:20px}
.scheda{background:#fff;padding:34px;border-radius:8px;width:100%;max-width:360px;box-shadow:0 12px 40px rgba(0,0,0,.28)}
h1{font-size:14px;letter-spacing:.14em;margin:0 0 4px;color:#1f4e78;font-weight:800}
h1 span{color:#e07a2f}
p{color:#6b7b8c;font-size:13px;margin:0 0 22px}
label{display:block;font-size:10px;letter-spacing:.10em;text-transform:uppercase;color:#6b7b8c;
 font-weight:700;margin:0 0 5px}
input{width:100%;padding:9px 10px;border:1px solid #dde2e8;border-radius:5px;font:inherit;margin-bottom:15px;box-sizing:border-box}
input:focus{outline:2px solid #1f4e78;outline-offset:-1px}
button{width:100%;background:#1f4e78;color:#fff;border:0;padding:11px;border-radius:5px;
 font-weight:700;font-size:14px;cursor:pointer}
button:hover{background:#143453}
.ko{background:#fdecea;color:#a3271a;padding:10px 12px;border-radius:5px;font-size:13px;margin-bottom:16px}
.campo-pw{position:relative}
.campo-pw input{padding-right:34px}
.campo-pw .occhio{position:absolute;right:4px;top:9px;background:none;border:0;cursor:pointer;
 font-size:15px;line-height:1;padding:6px;color:#6b7b8c}
.campo-pw .occhio:hover{color:#1f4e78}
</style></head><body>
<form class="scheda" method="post" action="/crm/accedi">
<h1>ENERGELIA<span>·</span>CRM</h1><p>Area riservata</p>
{% if errore %}<div class="ko">{{ errore }}</div>{% endif %}
<label for="email">Email</label><input id="email" name="email" type="email" required autofocus>
<label for="password">Password</label>
<div class="campo-pw"><input id="password" name="password" type="password" required>
<button type="button" class="occhio" onclick="mostraPw('password')" aria-label="Mostra password">occhio</button></div>
<button type="submit">Accedi</button>
</form>
<script>
function mostraPw(id){var c=document.getElementById(id);c.type=(c.type==='password')?'text':'password';}
</script>
</body></html>"""

T_CARICA_PUBBLICO = """<!doctype html><html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Carica documenti · Energelia</title>
<style>
body{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;
 background:#1f4e78;font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;padding:20px}
.scheda{background:#fff;padding:34px;border-radius:8px;width:100%;max-width:460px;box-shadow:0 12px 40px rgba(0,0,0,.28)}
h1{font-size:14px;letter-spacing:.14em;margin:0 0 4px;color:#1f4e78;font-weight:800}
h1 span{color:#e07a2f}
p{color:#6b7b8c;font-size:13px;margin:0 0 18px}
.nome-cliente{font-size:19px;font-weight:700;color:#1f4e78;margin:0 0 4px}
.zona{border:2px dashed #dde2e8;border-radius:8px;padding:26px 16px;text-align:center;margin-bottom:16px}
input[type=file]{width:100%;font-size:13px}
button{width:100%;background:#1f4e78;color:#fff;border:0;padding:11px;border-radius:5px;
 font-weight:700;font-size:14px;cursor:pointer}
button:hover{background:#143453}
.ok{background:#e8f5ee;border:1px solid #bfe0cd;color:#1c6340;padding:12px 14px;border-radius:5px;font-size:13px;margin-bottom:16px}
.ko{background:#fdecea;color:#a3271a;padding:10px 12px;border-radius:5px;font-size:13px;margin-bottom:16px}
</style></head><body>
<div class="scheda">
<h1>ENERGELIA<span>·</span>CRM</h1>
<p class="nome-cliente">{{ cliente.ragione_sociale }}</p>
{% if fatto is defined %}
  <div class="ok">{{ fatto }} file ricevuti, grazie. Puoi caricarne altri quando vuoi con lo stesso link.</div>
{% endif %}
{% if not configurato %}
  <div class="ko">Il caricamento non è ancora attivo. Mandaci i documenti direttamente via email, ci scusiamo per il disagio.</div>
{% else %}
  <p>Carica qui i documenti che ti abbiamo chiesto, anche in più volte e senza un ordine preciso — ci pensiamo noi a smistarli.</p>
  <form method="post" enctype="multipart/form-data">
    <div class="zona"><input type="file" name="file" multiple required></div>
    <button type="submit">Carica</button>
  </form>
{% endif %}
</div>
</body></html>"""

T_RICHIESTA_PUBBLICA = """<!doctype html><html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ p.nome_bando }} · Energelia</title>
<style>
body{margin:0;min-height:100vh;background:#f6f7f9;font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;padding:24px 16px}
.scheda{background:#fff;padding:30px;border-radius:8px;max-width:560px;margin:0 auto;box-shadow:0 8px 30px rgba(0,0,0,.08)}
h1{font-size:14px;letter-spacing:.14em;margin:0 0 4px;color:#1f4e78;font-weight:800}
h1 span{color:#e07a2f}
.nome-pratica{font-size:19px;font-weight:700;color:#1f4e78;margin:14px 0 2px}
.nome-cliente{color:#6b7b8c;font-size:13px;margin:0 0 20px}
.voce{border:1px solid #dde2e8;border-radius:6px;padding:14px 16px;margin-bottom:12px}
.voce label{display:block;font-weight:700;font-size:13px;color:#1f4e78;margin-bottom:8px}
.gia-ricevuto{font-size:12px;color:#1c6340;margin:-4px 0 8px}
input[type=text],textarea{width:100%;padding:8px 10px;border:1px solid #dde2e8;border-radius:5px;font:inherit;box-sizing:border-box}
textarea{min-height:70px;resize:vertical}
input[type=file]{width:100%;font-size:13px}
button{width:100%;background:#1f4e78;color:#fff;border:0;padding:11px;border-radius:5px;
 font-weight:700;font-size:14px;cursor:pointer;margin-top:6px}
button:hover{background:#143453}
.ok{background:#e8f5ee;border:1px solid #bfe0cd;color:#1c6340;padding:12px 14px;border-radius:5px;font-size:13px;margin-bottom:16px}
.ko{background:#fdecea;color:#a3271a;padding:10px 12px;border-radius:5px;font-size:13px;margin-bottom:16px}
</style></head><body>
<div class="scheda">
<h1>ENERGELIA<span>·</span>CRM</h1>
<p class="nome-pratica">{{ p.nome_bando }}</p>
<p class="nome-cliente">{{ p.cliente.ragione_sociale }}</p>
{% if fatto %}<div class="ok">Ricevuto, grazie. Puoi tornare su questo link in qualsiasi momento per completare o aggiungere altro.</div>{% endif %}
{% if not configurato %}
  <div class="ko">Il caricamento non è ancora attivo. Mandaci i documenti direttamente via email, ci scusiamo per il disagio.</div>
{% else %}
<form method="post" enctype="multipart/form-data">
{% for v in p.voci_richiesta %}
  <div class="voce">
    <label>{{ v.etichetta }}</label>
    {% if v.tipo_risposta == 'testo' %}
      <textarea name="voce_{{ v.id }}">{{ v.valore_testo or '' }}</textarea>
    {% else %}
      {% if v.compilata %}<p class="gia-ricevuto">Già ricevuto — puoi caricarne altri per questa voce.</p>{% endif %}
      <input type="file" name="voce_{{ v.id }}" multiple>
    {% endif %}
  </div>
{% endfor %}
  <div class="voce">
    <label>Altri documenti (facoltativo)</label>
    <input type="file" name="extra" multiple>
  </div>
  <button type="submit">Invia</button>
</form>
{% endif %}
</div>
</body></html>"""

T_DASHBOARD = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div>
  <h1>Dashboard</h1>
  <p class="sottotitolo">Situazione al {{ oggi }} · i numeri si aggiornano da soli a ogni modifica.</p>
</div><div><a class="btn" href="/crm/clienti/nuovo">Nuovo cliente</a>
<a class="btn ambra" href="/crm/pratiche/nuova">Nuova pratica</a></div></div>

<div class="griglia g4">
  <div class="kpi"><div class="etichetta">Clienti</div><div class="valore">{{ k.clienti }}</div>
    <div class="nota">{{ k.clienti_con_pratica }} con almeno una pratica</div></div>
  <div class="kpi"><div class="etichetta">Pratiche in lavorazione</div><div class="valore">{{ k.pratiche_aperte }}</div>
    <div class="nota">su {{ k.pratiche }} in totale</div></div>
  <div class="kpi"><div class="etichetta">Richiesto</div><div class="valore">{{ euro(k.richiesto) }}</div>
    <div class="nota">domande presentate o in corso</div></div>
  <div class="kpi"><div class="etichetta">Concesso</div><div class="valore">{{ euro(k.concesso) }}</div>
    <div class="nota">{{ k.ammesse }} {{ 'pratica' if k.ammesse == 1 else 'pratiche' }} con esito positivo</div></div>
</div>

<h2>Pipeline</h2>
<div class="pipeline">
{% for fase, breve, quanti in pipeline %}
  <a href="/crm/pratiche?fase={{ fase|urlencode }}" class="{{ 'vuota' if not quanti }}" title="{{ fase }}">
    <div class="num">{{ quanti }}</div><div class="nome">{{ breve }}</div>
    {% if quanti %}<div class="barra" style="width:{{ (quanti / massimo * 100)|round|int }}%"></div>{% endif %}
  </a>
{% endfor %}
</div>

<div class="griglia g2" style="margin-top:26px">
  <div>
    <h2 style="margin-top:0">Scadenze entro 60 giorni</h2>
    {% if scadenze %}
    <div class="tabella scorri"><table>
      <thead><tr><th>Pratica</th><th>Cliente</th><th>Scade</th><th>Mancano</th></tr></thead><tbody>
      {% for p in scadenze %}<tr>
        <td><a href="/crm/pratiche/{{ p.id }}">{{ p.nome_bando }}</a></td>
        <td>{{ p.cliente.ragione_sociale }}</td>
        <td class="num">{{ data_it(p.data_scadenza) }}</td>
        <td class="num {{ 'scad-rossa' if p.giorni_a_scadenza < 15 else 'scad-ambra' }}">
          {{ p.giorni_a_scadenza }} gg</td></tr>{% endfor %}
      </tbody></table></div>
    {% else %}<div class="vuoto">Nessuna scadenza nei prossimi 60 giorni.</div>{% endif %}

    <h2>Compensi Energelia</h2>
    <div class="griglia g3">
      <div class="kpi"><div class="etichetta">Corrispettivi</div>
        <div class="valore" style="font-size:20px">{{ euro(k.corrispettivi) }}</div></div>
      <div class="kpi"><div class="etichetta">Success fee maturata</div>
        <div class="valore" style="font-size:20px">{{ euro(k.success_fee) }}</div></div>
      <div class="kpi"><div class="etichetta">Incassato</div>
        <div class="valore" style="font-size:20px">{{ euro(k.incassato) }}</div></div>
    </div>
  </div>

  <div>
    <h2 style="margin-top:0">Ultime attività</h2>
    {% if attivita %}
    <ul class="diario">
    {% for a in attivita %}<li>
      <div class="capo"><span class="tipo">{{ a.tipo }}</span><span>{{ data_it(a.data) }}</span>
        <span>·</span><a href="/crm/clienti/{{ a.cliente_id }}">{{ a.cliente.ragione_sociale }}</a>
        {% if a.utente %}<span>· {{ a.utente.nome }}</span>{% endif %}</div>
      <div>{{ a.testo }}</div></li>{% endfor %}
    </ul>
    {% else %}<div class="vuoto">Nessuna attività registrata. Aprine una dalla scheda di un cliente.</div>{% endif %}
  </div>
</div>

<h2>Clienti da ricontattare</h2>
{% if fermi %}
<div class="tabella scorri"><table>
  <thead><tr><th>Cliente</th><th>Consulente</th><th>Ultimo contatto</th><th>Prossima azione</th></tr></thead><tbody>
  {% for c, giorni in fermi %}<tr>
    <td><a href="/crm/clienti/{{ c.id }}">{{ c.ragione_sociale }}</a></td>
    <td>{{ c.consulente.nome if c.consulente else '—' }}</td>
    <td class="num">{{ (giorni ~ ' giorni fa') if giorni is not none else 'mai' }}</td>
    <td>{{ c.prossima_azione or '—' }}</td></tr>{% endfor %}
</tbody></table></div>
{% else %}<div class="vuoto">Tutti i clienti sono stati sentiti di recente.</div>{% endif %}
{% endblock %}"""

T_CLIENTI = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>Clienti</h1>
<p class="sottotitolo">{{ elenco|length }} su {{ totale }} · l'elenco si aggiorna con i filtri qui sotto.</p></div>
<div><a class="btn chiaro" href="/crm/esporta.xlsx">Esporta in Excel</a>
<a class="btn" href="/crm/clienti/nuovo">Nuovo cliente</a></div></div>

<form class="filtri" method="get">
  <input name="q" value="{{ q or '' }}" placeholder="Cerca ragione sociale, P.IVA, referente, città" style="min-width:290px">
  <select name="canale"><option value="">Tutti i canali</option>
    {% for c in canali %}<option value="{{ c }}" {{ 'selected' if canale==c }}>{{ c }}</option>{% endfor %}</select>
  <select name="consulente"><option value="">Tutti i consulenti</option>
    {% for u in consulenti %}<option value="{{ u.id }}" {{ 'selected' if consulente==u.id|string }}>{{ u.nome }}</option>{% endfor %}</select>
  <button class="btn" type="submit">Filtra</button>
  {% if q or canale or consulente %}<a class="btn chiaro" href="/crm/clienti">Azzera</a>{% endif %}
</form>

{% if utente.is_admin %}
<details class="riquadro" style="margin-bottom:16px"><summary style="cursor:pointer;padding:14px 16px;font-weight:700;color:#1f4e78">Importa clienti da file (.xlsx o .csv)</summary>
<div class="corpo" style="padding-top:0">
  <p style="color:#6b7b8c">Il file .xlsx col foglio "Pratiche" importa anche le pratiche insieme ai clienti; un .csv importa solo i clienti.</p>
  <form method="post" action="/crm/importa" enctype="multipart/form-data" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <input type="file" name="file" accept=".xlsx,.csv" required>
    <button class="btn ambra" type="submit">Importa</button>
  </form>
</div></details>
{% endif %}

{% if elenco %}
<div class="tabella scorri"><table>
<thead><tr><th>Codice</th><th>Ragione sociale</th><th>Sede</th><th>Referente</th>
<th>Canale</th><th>Consulente</th><th>Pratiche</th><th>Ultimo contatto</th></tr></thead><tbody>
{% for c in elenco %}<tr>
  <td>{{ c.codice }}</td>
  <td><a href="/crm/clienti/{{ c.id }}"><strong>{{ c.ragione_sociale }}</strong></a>
      {% if c.dimensione %}<div style="font-size:11px;color:#6b7b8c">{{ c.dimensione }}{% if c.ateco %} · ATECO {{ c.ateco }}{% endif %}</div>{% endif %}</td>
  <td>{{ c.citta or '—' }}{% if c.provincia %} ({{ c.provincia }}){% endif %}</td>
  <td>{{ c.referente or '—' }}{% if c.telefono %}<div style="font-size:11px;color:#6b7b8c">{{ c.telefono }}</div>{% endif %}</td>
  <td>{{ c.canale or '—' }}</td>
  <td>{{ c.consulente.nome if c.consulente else '—' }}</td>
  <td class="num">{{ c.pratiche|length }}</td>
  <td class="num">{{ data_it(c.ultimo_contatto) }}</td>
</tr>{% endfor %}
</tbody></table></div>
{% else %}<div class="vuoto">Nessun cliente corrisponde ai filtri. <a href="/crm/clienti/nuovo">Aggiungine uno</a>.</div>{% endif %}
{% endblock %}"""

T_CLIENTE_FORM = """{% extends "base" %}{% block contenuto %}
<h1>{{ 'Modifica ' ~ cliente.ragione_sociale if cliente.id else 'Nuovo cliente' }}</h1>
<p class="sottotitolo">
  {{ cliente.codice or 'Il codice viene assegnato al salvataggio.' }}
  {% if lead %}· precompilato dal lead <strong>{{ lead.nome }}</strong>{% endif %}
  {% if da_visura %}· campi anagrafici presi dalla visura caricata{% endif %}
  {% if da_ricerca %}· contatti trovati con la ricerca online{% endif %}
</p>

{% if not cliente.id %}
<div class="riquadro" style="margin-bottom:20px"><div class="corpo">
  <h2 style="margin-top:0">Carica la visura camerale (facoltativo)</h2>
  {% if anthropic_ok %}
  <p style="color:#6b7b8c">Precompila ragione sociale, P.IVA, sede e legale rappresentante. Non tocca la
  composizione societaria: solo legale rappresentante ed eventuali titolari effettivi.</p>
  <form method="post" action="/crm/clienti/nuovo-da-visura" enctype="multipart/form-data" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    {% if lead %}<input type="hidden" name="lead_id" value="{{ lead.id }}">{% endif %}
    <input type="file" name="file" accept=".pdf" required>
    <button class="btn ambra" type="submit">Estrai dalla visura</button>
  </form>
  {% else %}
  <p style="color:#6b7b8c">Richiede ANTHROPIC_API_KEY, non ancora configurata.</p>
  {% endif %}
</div></div>
{% endif %}

{% if cliente.id %}
<div class="riquadro" style="margin-bottom:20px"><div class="corpo">
  <h2 style="margin-top:0">Cerca informazioni online</h2>
  {% if anthropic_ok %}
  <p style="color:#6b7b8c">Cerca sul web telefono, email, PEC e indirizzo mancanti — non tocca i campi già
  compilati. Controlla sempre prima di salvare.</p>
  <form method="post" action="/crm/clienti/{{ cliente.id }}/cerca-online">
    <button class="btn ambra" type="submit">Cerca online</button>
  </form>
  {% else %}
  <p style="color:#6b7b8c">Richiede ANTHROPIC_API_KEY, non ancora configurata.</p>
  {% endif %}
</div></div>
{% endif %}

<form method="post" action="{{ ('/crm/clienti/' ~ cliente.id ~ '/modifica') if cliente.id else '/crm/clienti/nuovo' }}">
{% if lead %}<input type="hidden" name="lead_id" value="{{ lead.id }}">{% endif %}
<fieldset><legend>Azienda</legend><div class="griglia g3">
  <label><span class="etichetta">Ragione sociale *</span><input name="ragione_sociale" required value="{{ cliente.ragione_sociale or '' }}"></label>
  <label><span class="etichetta">Partita IVA / CF</span><input name="piva" value="{{ cliente.piva or '' }}"></label>
  <label><span class="etichetta">Settore ATECO</span><input name="ateco" value="{{ cliente.ateco or '' }}"></label>
  <label><span class="etichetta">Dimensione impresa</span><select name="dimensione"><option value=""></option>
    {% for x in dimensioni %}<option {{ 'selected' if cliente.dimensione==x }}>{{ x }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Città sede legale</span><input name="citta" value="{{ cliente.citta or '' }}"></label>
  <label><span class="etichetta">Provincia</span><input name="provincia" maxlength="4" value="{{ cliente.provincia or '' }}"></label>
  <label><span class="etichetta">Regione</span><input name="regione" value="{{ cliente.regione or '' }}"></label>
</div></fieldset>
<fieldset><legend>Contatti</legend><div class="griglia g3">
  <label><span class="etichetta">Referente</span><input name="referente" value="{{ cliente.referente or '' }}"></label>
  <label><span class="etichetta">Ruolo referente</span><input name="ruolo_referente" value="{{ cliente.ruolo_referente or '' }}"></label>
  <label><span class="etichetta">Telefono</span><input name="telefono" value="{{ cliente.telefono or '' }}"></label>
  <label><span class="etichetta">Email</span><input name="email" type="email" value="{{ cliente.email or '' }}"></label>
  <label><span class="etichetta">PEC</span><input name="pec" value="{{ cliente.pec or '' }}"></label>
</div></fieldset>
<fieldset><legend>Dati economici</legend><div class="griglia g3">
  <label><span class="etichetta">Codice Fiscale</span><input name="codice_fiscale" value="{{ cliente.codice_fiscale or '' }}"></label>
  <label><span class="etichetta">Codice destinatario SDI</span><input name="codice_sdi" value="{{ cliente.codice_sdi or '' }}"></label>
  <label><span class="etichetta">Intestatario conto (se diverso)</span><input name="intestatario_conto" value="{{ cliente.intestatario_conto or '' }}"></label>
  <label><span class="etichetta">IBAN del cliente</span><input name="iban" value="{{ cliente.iban or '' }}" placeholder="Dove riceve l'accredito del bando"></label>
  <label><span class="etichetta">Email di fatturazione</span><input name="email_fatturazione" type="email" value="{{ cliente.email_fatturazione or '' }}"></label>
  <label><span class="etichetta">Titolari effettivi</span><input name="titolari_effettivi" value="{{ cliente.titolari_effettivi or '' }}" placeholder="Un nome per riga"></label>
</div></fieldset>
<fieldset><legend>Gestione</legend><div class="griglia g3">
  <label><span class="etichetta">Consulente Energelia</span><select name="consulente_id"><option value=""></option>
    {% for u in consulenti %}<option value="{{ u.id }}" {{ 'selected' if cliente.consulente_id==u.id }}>{{ u.nome }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Canale acquisizione</span><select name="canale"><option value=""></option>
    {% for x in canali %}<option {{ 'selected' if cliente.canale==x }}>{{ x }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Data primo contatto</span><input name="data_primo_contatto" type="date" value="{{ cliente.data_primo_contatto or '' }}"></label>
  <label><span class="etichetta">Prossima azione</span><input name="prossima_azione" value="{{ cliente.prossima_azione or '' }}"></label>
</div>
<label><span class="etichetta">Note</span><textarea name="note">{{ cliente.note or '' }}</textarea></label>
</fieldset>
<button class="btn" type="submit">Salva cliente</button>
<a class="btn chiaro" href="{{ '/crm/clienti/' ~ cliente.id if cliente.id else '/crm/clienti' }}">Annulla</a>
</form>{% endblock %}"""

T_CLIENTE = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>{{ c.ragione_sociale }}</h1>
<p class="sottotitolo">{{ c.codice }}{% if c.citta %} · {{ c.citta }}{% if c.provincia %} ({{ c.provincia }}){% endif %}{% endif %}
{% if c.dimensione %} · {{ c.dimensione }} impresa{% endif %}</p></div>
<div><a class="btn chiaro" href="/crm/clienti/{{ c.id }}/modifica">Modifica</a>
<a class="btn" href="/crm/pratiche/nuova?cliente={{ c.id }}">Nuova pratica</a></div></div>

<div class="griglia g2">
<div><dl class="dettaglio">
  <dt>Partita IVA</dt><dd>{{ c.piva or '—' }}</dd>
  <dt>ATECO</dt><dd>{{ c.ateco or '—' }}</dd>
  <dt>Referente</dt><dd>{{ c.referente or '—' }}{% if c.ruolo_referente %} · {{ c.ruolo_referente }}{% endif %}</dd>
  <dt>Telefono</dt><dd>{{ c.telefono or '—' }}</dd>
  <dt>Email</dt><dd>{% if c.email %}<a href="mailto:{{ c.email }}">{{ c.email }}</a>{% else %}—{% endif %}</dd>
  <dt>PEC</dt><dd>{{ c.pec or '—' }}</dd>
  <dt>Codice Fiscale</dt><dd>{{ c.codice_fiscale or '—' }}</dd>
  <dt>Codice SDI</dt><dd>{{ c.codice_sdi or '—' }}</dd>
  <dt>IBAN cliente</dt><dd>{{ c.iban or '—' }}</dd>
  <dt>Intestatario conto</dt><dd>{{ c.intestatario_conto or '—' }}</dd>
  <dt>Email fatturazione</dt><dd>{{ c.email_fatturazione or '—' }}</dd>
  <dt>Titolari effettivi</dt><dd style="white-space:pre-wrap">{{ c.titolari_effettivi or '—' }}</dd>
  <dt>Consulente</dt><dd>{{ c.consulente.nome if c.consulente else '—' }}</dd>
  <dt>Canale</dt><dd>{{ c.canale or '—' }}</dd>
  <dt>Primo contatto</dt><dd>{{ data_it(c.data_primo_contatto) }}</dd>
  <dt>Prossima azione</dt><dd>{{ c.prossima_azione or '—' }}</dd>
  <dt>Note</dt><dd>{{ c.note or '—' }}</dd>
</dl></div>

<div>
<h2 style="margin-top:0">Registra un'attività</h2>
<form method="post" action="/crm/attivita/nuova" class="riquadro">
  <div class="corpo">
  <input type="hidden" name="cliente_id" value="{{ c.id }}">
  <div class="griglia g3">
    <label><span class="etichetta">Data</span><input name="data" type="date" value="{{ oggi_iso }}"></label>
    <label><span class="etichetta">Tipo</span><select name="tipo">
      {% for t in tipi %}<option>{{ t }}</option>{% endfor %}</select></label>
    <label><span class="etichetta">Pratica</span><select name="pratica_id"><option value="">—</option>
      {% for p in c.pratiche %}<option value="{{ p.id }}">{{ p.nome_bando }}</option>{% endfor %}</select></label>
  </div>
  <label><span class="etichetta">Cosa è successo</span><textarea name="testo" required
    placeholder="Es. Sentito il titolare: conferma di voler procedere, manda visura entro venerdì."></textarea></label>
  <button class="btn" type="submit">Registra</button>
  </div>
</form>
</div>
</div>

<h2>Documenti dal cliente</h2>
<div class="riquadro"><div class="corpo">
  <p style="margin-top:0;color:#6b7b8c">Manda questo link al cliente: può caricare documenti in autonomia,
  senza login, anche in più volte.</p>
  <div class="griglia g2">
    <input readonly id="link-caricamento" value="{{ request.url_root.rstrip('/') }}/crm/carica/{{ c.token_caricamento }}"
      style="font-size:12px">
    <button type="button" class="btn chiaro" onclick="
      navigator.clipboard.writeText(document.getElementById('link-caricamento').value);
      this.textContent='Copiato!'; setTimeout(()=>this.textContent='Copia link', 1500);
    ">Copia link</button>
  </div>
  <form method="post" action="/crm/clienti/{{ c.id }}/invia-link" class="griglia g2" style="margin-top:10px">
    <input type="email" name="destinatario" required placeholder="Email del cliente" value="{{ c.email or '' }}">
    <button class="btn ambra" type="submit">Invia via email</button>
  </form>
</div></div>

{% if documenti %}
<div class="tabella scorri"><table>
<thead><tr><th>File</th><th>Caricato</th><th>Stato</th><th></th></tr></thead><tbody>
{% for doc in documenti %}<tr>
  <td>{% if doc.link_drive %}<a href="{{ doc.link_drive }}" target="_blank">{{ doc.nome_file }}</a>
      {% else %}{{ doc.nome_file }}{% endif %}</td>
  <td>{{ data_it(doc.creato_il.date()) }} · {{ doc.caricato_da }}</td>
  <td>{% if doc.stato == 'assegnato' %}<span class="pill ok">→ {{ doc.pratica.nome_bando if doc.pratica else 'assegnato' }}</span>
      {% else %}<span class="pill">da smistare</span>{% endif %}</td>
  <td class="num">
    {% if doc.stato != 'assegnato' and c.pratiche %}
    <form method="post" action="/crm/documenti/{{ doc.id }}/assegna" style="display:inline-flex;gap:6px">
      <select name="pratica_id" required><option value="">Assegna a…</option>
        {% for p in c.pratiche %}<option value="{{ p.id }}">{{ p.nome_bando }}</option>{% endfor %}</select>
      <button class="btn chiaro" type="submit">Ok</button>
    </form>{% endif %}
  </td>
</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="vuoto">Nessun documento caricato finora.</div>{% endif %}

<h2>Pratiche ({{ c.pratiche|length }})</h2>
{% if c.pratiche %}
<div class="tabella scorri"><table>
<thead><tr><th>Codice</th><th>Bando</th><th>Fase</th><th>Priorità</th><th>Scadenza</th>
<th>Richiesto</th><th>Concesso</th></tr></thead><tbody>
{% for p in c.pratiche %}<tr>
  <td>{{ p.codice }}</td><td><a href="/crm/pratiche/{{ p.id }}">{{ p.nome_bando }}</a></td>
  <td><span class="pill {{ classe_fase(p.fase) }}">{{ p.fase }}</span></td>
  <td><span class="pill p-{{ (p.priorita or 'media')|lower }}">{{ p.priorita }}</span></td>
  <td class="num">{{ data_it(p.data_scadenza) }}</td>
  <td class="num">{{ euro(p.importo_richiesto) }}</td>
  <td class="num">{{ euro(p.importo_concesso) }}</td>
</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="vuoto">Nessuna pratica. <a href="/crm/pratiche/nuova?cliente={{ c.id }}">Aprine una</a>.</div>{% endif %}

<h2>Diario ({{ c.attivita|length }})</h2>
{% if c.attivita %}<ul class="diario">
{% for a in c.attivita %}<li>
  <div class="capo"><span class="tipo">{{ a.tipo }}</span><span>{{ data_it(a.data) }}</span>
  {% if a.utente %}<span>· {{ a.utente.nome }}</span>{% endif %}
  {% if a.pratica %}<span>· {{ a.pratica.nome_bando }}</span>{% endif %}</div>
  <div>{{ a.testo }}</div></li>{% endfor %}
</ul>{% else %}<div class="vuoto">Il diario è vuoto. La prima nota che registri finisce qui.</div>{% endif %}
{% endblock %}"""

T_PRATICHE = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>Pratiche</h1>
<p class="sottotitolo">{{ elenco|length }} su {{ totale }} · richiesto {{ euro(tot_richiesto) }} · concesso {{ euro(tot_concesso) }}</p></div>
<div><a class="btn chiaro" href="/crm/esporta.xlsx">Esporta in Excel</a>
<a class="btn" href="/crm/pratiche/nuova">Nuova pratica</a></div></div>

<form class="filtri" method="get">
  <input name="q" value="{{ q or '' }}" placeholder="Cerca bando, ente, cliente" style="min-width:250px">
  <select name="fase"><option value="">Tutte le fasi</option>
    {% for f in fasi %}<option value="{{ f }}" {{ 'selected' if fase==f }}>{{ f }}</option>{% endfor %}</select>
  <select name="priorita"><option value="">Tutte le priorità</option>
    {% for p in priorita_lista %}<option value="{{ p }}" {{ 'selected' if priorita==p }}>{{ p }}</option>{% endfor %}</select>
  <label style="margin:0;display:flex;align-items:center;gap:6px;font-size:13px">
    <input type="checkbox" name="aperte" value="1" {{ 'checked' if aperte }} style="width:auto"> Solo in lavorazione</label>
  <button class="btn" type="submit">Filtra</button>
  {% if q or fase or priorita or aperte %}<a class="btn chiaro" href="/crm/pratiche">Azzera</a>{% endif %}
</form>

{% if elenco %}
<div class="tabella scorri"><table>
<thead><tr><th>Codice</th><th>Bando</th><th>Cliente</th><th>Fase</th><th>Pri.</th>
<th>Scadenza</th><th>Richiesto</th><th>Concesso</th><th>Compenso</th><th>Prossimo step</th></tr></thead><tbody>
{% for p in elenco %}<tr>
  <td>{{ p.codice }}</td>
  <td><a href="/crm/pratiche/{{ p.id }}"><strong>{{ p.nome_bando }}</strong></a>
      {% if p.ente %}<div style="font-size:11px;color:#6b7b8c">{{ p.ente }}</div>{% endif %}</td>
  <td><a href="/crm/clienti/{{ p.cliente_id }}">{{ p.cliente.ragione_sociale }}</a></td>
  <td><span class="pill {{ classe_fase(p.fase) }}">{{ p.fase }}</span></td>
  <td><span class="pill p-{{ (p.priorita or 'media')|lower }}">{{ p.priorita }}</span></td>
  <td class="num">{{ data_it(p.data_scadenza) }}
    {% set g = p.giorni_a_scadenza %}
    {% if g is not none and g < 60 %}<div class="{{ 'scad-rossa' if g < 15 else 'scad-ambra' }}"
      style="font-size:11px">{{ g }} gg</div>{% endif %}</td>
  <td class="num">{{ euro(p.importo_richiesto) }}</td>
  <td class="num">{{ euro(p.importo_concesso) }}</td>
  <td class="num">{{ euro(p.compenso_totale) }}</td>
  <td>{{ p.prossimo_step or '—' }}</td>
</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="vuoto">Nessuna pratica corrisponde ai filtri.</div>{% endif %}
{% endblock %}"""

T_PRATICA_FORM = """{% extends "base" %}{% block contenuto %}
<h1>{{ 'Modifica pratica' if p.id else 'Nuova pratica' }}</h1>
<p class="sottotitolo">
  {{ p.codice or 'Il codice viene assegnato al salvataggio.' }}
  {% if p.bando_id %}· precompilata dal bando <a href="/crm/bandi/{{ p.bando_id }}">{{ p.bando.nome if p.bando else '' }}</a>{% endif %}
</p>
{% if not p.id and bandi %}
<div class="riquadro" style="margin-bottom:16px"><div class="corpo">
  <label><span class="etichetta">Precompila da un bando salvato (facoltativo)</span>
  <select onchange="
    var parametri = new URLSearchParams(window.location.search);
    if (this.value) { parametri.set('bando', this.value); } else { parametri.delete('bando'); }
    window.location = '/crm/pratiche/nuova?' + parametri.toString();
  "><option value="">— compila a mano —</option>
    {% for bd in bandi %}<option value="{{ bd.id }}" {{ 'selected' if p.bando_id==bd.id }}>{{ bd.nome }}</option>{% endfor %}
  </select></label>
</div></div>
{% endif %}
<form method="post">
{% if p.bando_id %}<input type="hidden" name="bando_id" value="{{ p.bando_id }}">{% endif %}
<fieldset><legend>Bando</legend><div class="griglia g3">
  <label><span class="etichetta">Cliente *</span><select name="cliente_id" required><option value=""></option>
    {% for c in clienti %}<option value="{{ c.id }}" {{ 'selected' if p.cliente_id==c.id }}>{{ c.ragione_sociale }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Nome bando *</span><input name="nome_bando" required value="{{ p.nome_bando or '' }}"></label>
  <label><span class="etichetta">Ente erogatore</span><input name="ente" value="{{ p.ente or '' }}"></label>
  <label><span class="etichetta">Tipologia</span><select name="tipologia"><option value=""></option>
    {% for x in tipologie %}<option {{ 'selected' if p.tipologia==x }}>{{ x }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">% contributo</span><input name="perc_contributo" inputmode="decimal"
    placeholder="75" value="{{ p.perc_contributo or '' }}"></label>
  <label><span class="etichetta">Importo massimo erogabile</span><input name="importo_max"
    placeholder="30.000 linea A / 40.000 linea B" value="{{ p.importo_max or '' }}"></label>
  <label><span class="etichetta">Apertura bando</span><input name="data_apertura" type="date" value="{{ p.data_apertura or '' }}"></label>
  <label><span class="etichetta">Scadenza presentazione</span><input name="data_scadenza" type="date" value="{{ p.data_scadenza or '' }}"></label>
</div></fieldset>
<fieldset><legend>Avanzamento</legend><div class="griglia g3">
  <label><span class="etichetta">Fase attuale</span><select name="fase">
    {% for x in fasi %}<option {{ 'selected' if p.fase==x }}>{{ x }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Priorità</span><select name="priorita">
    {% for x in priorita_lista %}<option {{ 'selected' if p.priorita==x }}>{{ x }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Prossimo step</span><input name="prossimo_step" value="{{ p.prossimo_step or '' }}"></label>
  <label><span class="etichetta">Data presentazione</span><input name="data_presentazione" type="date" value="{{ p.data_presentazione or '' }}"></label>
  <label><span class="etichetta">Data esito</span><input name="data_esito" type="date" value="{{ p.data_esito or '' }}"></label>
  <label><span class="etichetta">Scadenza rendicontazione</span><input name="scadenza_rendicontazione" type="date" value="{{ p.scadenza_rendicontazione or '' }}"></label>
  <label><span class="etichetta">Importo richiesto €</span><input name="importo_richiesto" inputmode="decimal" value="{{ p.importo_richiesto or '' }}"></label>
  <label><span class="etichetta">Importo concesso €</span><input name="importo_concesso" inputmode="decimal" value="{{ p.importo_concesso or '' }}"></label>
</div>
<label><span class="etichetta">Documenti mancanti</span><textarea name="documenti_mancanti">{{ p.documenti_mancanti or '' }}</textarea></label>
</fieldset>
<fieldset><legend>Compenso Energelia</legend><div class="griglia g4">
  <label><span class="etichetta">Corrispettivo €</span><input name="corrispettivo" inputmode="decimal" value="{{ p.corrispettivo or '' }}"></label>
  <label><span class="etichetta">Success fee %</span><input name="success_fee_perc" inputmode="decimal" placeholder="5" value="{{ p.success_fee_perc or '' }}"></label>
  <label><span class="etichetta">Fatturato €</span><input name="fatturato" inputmode="decimal" value="{{ p.fatturato or '' }}"></label>
  <label><span class="etichetta">Incassato €</span><input name="incassato" inputmode="decimal" value="{{ p.incassato or '' }}"></label>
  <label><span class="etichetta">Conto d'incasso Energelia</span><select name="conto_incasso_id"><option value="">—</option>
    {% for conto in conti %}<option value="{{ conto.id }}" {{ 'selected' if p.conto_incasso_id==conto.id }}>{{ conto.nome }}</option>{% endfor %}</select></label>
</div>
<label><span class="etichetta">Note</span><textarea name="note">{{ p.note or '' }}</textarea></label>
</fieldset>
<button class="btn" type="submit">Salva pratica</button>
<a class="btn chiaro" href="{{ '/crm/pratiche/' ~ p.id if p.id else '/crm/pratiche' }}">Annulla</a>
</form>{% endblock %}"""

T_PRATICA = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>{{ p.nome_bando }}</h1>
<p class="sottotitolo">{{ p.codice }} · <a href="/crm/clienti/{{ p.cliente_id }}">{{ p.cliente.ragione_sociale }}</a>
{% if p.ente %} · {{ p.ente }}{% endif %}</p></div>
<div><a class="btn chiaro" href="/crm/pratiche/{{ p.id }}/modifica">Modifica</a></div></div>

<div class="griglia g4" style="margin-bottom:20px">
  <div class="kpi"><div class="etichetta">Fase</div>
    <div style="margin-top:9px"><span class="pill {{ classe_fase(p.fase) }}">{{ p.fase }}</span></div></div>
  <div class="kpi"><div class="etichetta">Scadenza</div>
    <div class="valore" style="font-size:19px">{{ data_it(p.data_scadenza) }}</div>
    {% set g = p.giorni_a_scadenza %}
    {% if g is not none %}<div class="nota {{ 'scad-rossa' if g < 15 else ('scad-ambra' if g < 60 else '') }}">
      {{ 'scaduta da ' ~ (-g) ~ ' giorni' if g < 0 else 'mancano ' ~ g ~ ' giorni' }}</div>{% endif %}</div>
  <div class="kpi"><div class="etichetta">Richiesto</div>
    <div class="valore" style="font-size:19px">{{ euro(p.importo_richiesto) }}</div></div>
  <div class="kpi"><div class="etichetta">Concesso</div>
    <div class="valore" style="font-size:19px">{{ euro(p.importo_concesso) }}</div></div>
</div>

<div class="griglia g2">
<div><h2 style="margin-top:0">Bando</h2><dl class="dettaglio">
  <dt>Tipologia</dt><dd>{{ p.tipologia or '—' }}</dd>
  <dt>% contributo</dt><dd>{{ (p.perc_contributo|string ~ ' %') if p.perc_contributo else '—' }}</dd>
  <dt>Importo massimo</dt><dd>{{ p.importo_max or '—' }}</dd>
  <dt>Apertura</dt><dd>{{ data_it(p.data_apertura) }}</dd>
  <dt>Presentazione</dt><dd>{{ data_it(p.data_presentazione) }}</dd>
  <dt>Esito</dt><dd>{{ data_it(p.data_esito) }}</dd>
  <dt>Rendicontazione</dt><dd>{{ data_it(p.scadenza_rendicontazione) }}</dd>
  <dt>Priorità</dt><dd><span class="pill p-{{ (p.priorita or 'media')|lower }}">{{ p.priorita }}</span></dd>
  <dt>Prossimo step</dt><dd>{{ p.prossimo_step or '—' }}</dd>
  <dt>Documenti mancanti</dt><dd>{{ p.documenti_mancanti or '—' }}</dd>
  <dt>Note</dt><dd>{{ p.note or '—' }}</dd>
</dl></div>
<div><h2 style="margin-top:0">Compenso Energelia</h2><dl class="dettaglio">
  <dt>Corrispettivo</dt><dd>{{ euro(p.corrispettivo) }}</dd>
  <dt>Success fee</dt><dd>{{ (p.success_fee_perc|string ~ ' %') if p.success_fee_perc else '—' }}</dd>
  <dt>Success fee maturata</dt><dd>{{ euro(p.success_fee_maturata) }}</dd>
  <dt>Totale</dt><dd><strong>{{ euro(p.compenso_totale) }}</strong></dd>
  <dt>Fatturato</dt><dd>{{ euro(p.fatturato) }}</dd>
  <dt>Incassato</dt><dd>{{ euro(p.incassato) }}</dd>
  <dt>Conto d'incasso</dt><dd>{{ p.conto_incasso.nome if p.conto_incasso else '—' }}</dd>
</dl>

<h2>Registra un'attività</h2>
<form method="post" action="/crm/attivita/nuova" class="riquadro"><div class="corpo">
  <input type="hidden" name="cliente_id" value="{{ p.cliente_id }}">
  <input type="hidden" name="pratica_id" value="{{ p.id }}">
  <div class="griglia g2">
    <label><span class="etichetta">Data</span><input name="data" type="date" value="{{ oggi_iso }}"></label>
    <label><span class="etichetta">Tipo</span><select name="tipo">
      {% for t in tipi %}<option>{{ t }}</option>{% endfor %}</select></label>
  </div>
  <label><span class="etichetta">Cosa è successo</span><textarea name="testo" required></textarea></label>
  <button class="btn" type="submit">Registra</button>
</div></form>
</div></div>

<h2>Diario della pratica</h2>
{% if p.attivita %}<ul class="diario">
{% for a in p.attivita|sort(attribute='data', reverse=true) %}<li>
  <div class="capo"><span class="tipo">{{ a.tipo }}</span><span>{{ data_it(a.data) }}</span>
  {% if a.utente %}<span>· {{ a.utente.nome }}</span>{% endif %}</div>
  <div>{{ a.testo }}</div></li>{% endfor %}
</ul>{% else %}<div class="vuoto">Nessuna attività su questa pratica.</div>{% endif %}

{% if p.bando and p.bando.approfondimento %}
<h2>Approfondimento sul bando</h2>
<div class="riquadro"><div class="corpo" style="white-space:pre-wrap">{{ p.bando.approfondimento }}</div></div>
{% endif %}

<h2>Richiesta al cliente per questa pratica</h2>
<div class="riquadro"><div class="corpo">
  <p style="margin-top:0;color:#6b7b8c">Un link diverso da quello generico del cliente: mostra solo le voci
  che elenchi qui sotto, ognuna con il suo spazio dedicato.</p>
  <div class="griglia g2">
    <input readonly id="link-pratica" value="{{ request.url_root.rstrip('/') }}/crm/richiesta/{{ p.token_caricamento }}"
      style="font-size:12px">
    <button type="button" class="btn chiaro" onclick="
      navigator.clipboard.writeText(document.getElementById('link-pratica').value);
      this.textContent='Copiato!'; setTimeout(()=>this.textContent='Copia link', 1500);
    ">Copia link</button>
  </div>
  <form method="post" action="/crm/pratiche/{{ p.id }}/invia-link" class="griglia g2" style="margin-top:10px">
    <input type="email" name="destinatario" required placeholder="Email del cliente" value="{{ p.cliente.email or '' }}">
    <button class="btn ambra" type="submit">Invia via email</button>
  </form>
</div></div>

{% if p.voci_richiesta %}
<div class="tabella scorri"><table>
<thead><tr><th>Voce</th><th>Tipo</th><th>Stato</th><th></th></tr></thead><tbody>
{% for v in p.voci_richiesta %}<tr>
  <td>{{ v.etichetta }}</td>
  <td>{{ 'Caricamento file' if v.tipo_risposta == 'file' else 'Testo libero' }}</td>
  <td>{% if v.compilata %}
        {% if v.tipo_risposta == 'file' and v.documento and v.documento.link_drive %}
          <a href="{{ v.documento.link_drive }}" target="_blank">✓ ricevuto</a>
        {% elif v.tipo_risposta == 'testo' %}<span class="pill ok">✓ {{ v.valore_testo[:60] }}{{ '…' if v.valore_testo and v.valore_testo|length > 60 }}</span>
        {% else %}<span class="pill ok">✓ ricevuto</span>{% endif %}
      {% else %}<span class="pill">in attesa</span>{% endif %}</td>
  <td class="num"><form method="post" action="/crm/voci/{{ v.id }}/elimina" style="display:inline">
    <button class="btn chiaro" type="submit">Rimuovi</button></form></td>
</tr>{% endfor %}</tbody></table></div>
{% endif %}

<form method="post" action="/crm/pratiche/{{ p.id }}/voci/nuova" class="riquadro"><div class="corpo">
  <div class="griglia g3">
    <label><span class="etichetta">Etichetta</span><input name="etichetta" required placeholder="Es. Preventivo fornitore"></label>
    <label><span class="etichetta">Tipo di risposta</span><select name="tipo_risposta">
      <option value="file">Caricamento file</option><option value="testo">Testo libero</option></select></label>
  </div>
  <button class="btn ambra" type="submit">Aggiungi voce</button>
</div></form>
{% endblock %}"""

T_ATTIVITA = """{% extends "base" %}{% block contenuto %}
<h1>Attività</h1><p class="sottotitolo">Ultime {{ elenco|length }} registrazioni di tutto il gruppo.</p>
<form class="filtri" method="get">
  <input name="q" value="{{ q or '' }}" placeholder="Cerca nel testo o nel cliente" style="min-width:280px">
  <select name="utente"><option value="">Tutti</option>
    {% for u in utenti %}<option value="{{ u.id }}" {{ 'selected' if utente_id==u.id|string }}>{{ u.nome }}</option>{% endfor %}</select>
  <button class="btn" type="submit">Filtra</button>
  {% if q or utente_id %}<a class="btn chiaro" href="/crm/attivita">Azzera</a>{% endif %}
</form>
{% if elenco %}<ul class="diario">
{% for a in elenco %}<li>
  <div class="capo"><span class="tipo">{{ a.tipo }}</span><span>{{ data_it(a.data) }}</span>
  <span>·</span><a href="/crm/clienti/{{ a.cliente_id }}">{{ a.cliente.ragione_sociale }}</a>
  {% if a.pratica %}<span>· <a href="/crm/pratiche/{{ a.pratica_id }}">{{ a.pratica.nome_bando }}</a></span>{% endif %}
  {% if a.utente %}<span>· {{ a.utente.nome }}</span>{% endif %}</div>
  <div>{{ a.testo }}</div></li>{% endfor %}
</ul>{% else %}<div class="vuoto">Nessuna attività registrata.</div>{% endif %}
{% endblock %}"""

T_LEAD = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>Lead</h1>
<p class="sottotitolo">{{ totale }} nominativi{% if q or stato or fonte %} (filtrati){% endif %} · dagli scraper</p></div>
<div><a class="btn" href="/crm/clienti/nuovo">Nuovo cliente da zero</a></div></div>

<form class="filtri" method="get">
  <input name="q" value="{{ q or '' }}" placeholder="Cerca nome, email, telefono" style="min-width:240px">
  <select name="stato"><option value="">Tutti gli stati</option>
    {% for x in stati_lead %}<option {{ 'selected' if stato==x }}>{{ x }}</option>{% endfor %}</select>
  <select name="fonte"><option value="">Tutte le fonti</option>
    {% for x in fonti %}<option {{ 'selected' if fonte==x }}>{{ x }}</option>{% endfor %}</select>
  <select name="provincia"><option value="">Tutte le province</option>
    {% for x in province %}<option {{ 'selected' if provincia==x }}>{{ x }}</option>{% endfor %}</select>
  <button class="btn" type="submit">Filtra</button>
  {% if q or stato or fonte %}<a class="btn chiaro" href="/crm/lead">Azzera</a>{% endif %}
</form>

{% if utente.is_admin %}
<details class="riquadro" style="margin-bottom:16px"><summary style="cursor:pointer;padding:14px 16px;font-weight:700;color:#1f4e78">Importa lead da scraping (.xlsx o .csv)</summary>
<div class="corpo" style="padding-top:0">
  <p style="color:#6b7b8c">Carica l'output di uno scraper così com'è: riconosce da sé le colonne e salta le righe separatore.</p>
  <form method="post" action="/crm/lead/importa" enctype="multipart/form-data" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <input name="fonte" placeholder="Etichetta di questo giro, es. Umbria Energia" style="min-width:240px">
    <input type="file" name="file" accept=".xlsx,.csv" required>
    <button class="btn ambra" type="submit">Importa</button>
  </form>
</div></details>
{% endif %}

{% if elenco %}
<div class="tabella scorri"><table>
<thead><tr><th>Nome</th><th>Tipo</th><th>Contatto</th><th>Fonte</th><th>Stato</th><th></th></tr></thead><tbody>
{% for l in elenco %}<tr>
  <td>{{ l.nome }}{% if l.indirizzo %}<div class="nota">{{ l.indirizzo }}</div>
    {% elif l.comune or l.provincia %}<div class="nota">{{ [l.comune, l.provincia]|select|join(' · ') }}</div>{% endif %}</td>
  <td>{{ l.tipo or '—' }}</td>
  <td>{{ l.contatto_migliore }}</td>
  <td>{{ l.fonte or '—' }}</td>
  <td><span class="pill {{ 'ok' if l.stato=='convertito' else ('ko' if l.stato=='scartato' else '') }}">{{ l.stato }}</span></td>
  <td class="num">
    {% if l.stato == 'convertito' %}<a class="btn chiaro" href="/crm/clienti/{{ l.cliente_id }}">Vedi cliente</a>
    {% else %}
    <a class="btn chiaro" href="/crm/clienti/nuovo?lead={{ l.id }}">Converti</a>
    <form method="post" action="/crm/lead/{{ l.id }}/stato" style="display:inline">
      <input type="hidden" name="stato" value="{{ 'scartato' if l.stato != 'scartato' else 'nuovo' }}">
      <button class="btn chiaro" type="submit">{{ 'Scarta' if l.stato != 'scartato' else 'Ripristina' }}</button>
    </form>
    {% endif %}
  </td>
</tr>{% endfor %}</tbody></table></div>
<div class="filtri" style="margin-top:14px">
  {% if pagina_n > 1 %}<a class="btn chiaro" href="?{{ query_senza_pagina }}&pagina={{ pagina_n - 1 }}">« Precedenti</a>{% endif %}
  <span class="nota">Pagina {{ pagina_n }} di {{ pagine_tot }}</span>
  {% if pagina_n < pagine_tot %}<a class="btn chiaro" href="?{{ query_senza_pagina }}&pagina={{ pagina_n + 1 }}">Successivi »</a>{% endif %}
</div>
{% else %}<div class="vuoto">Nessun lead corrisponde ai filtri.</div>{% endif %}
{% endblock %}"""

T_DOCUMENTI = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>Documenti</h1>
<p class="sottotitolo">{{ elenco|length }} file{{ ' da smistare' if solo_da_smistare else '' }}</p></div>
<div>{% if solo_da_smistare %}<a class="btn chiaro" href="/crm/documenti?tutti=1">Mostra tutti</a>
{% else %}<a class="btn chiaro" href="/crm/documenti">Solo da smistare</a>{% endif %}</div></div>

{% if not configurato %}
<div class="vuoto">Il caricamento da Google Drive non è ancora configurato.</div>
{% elif elenco %}
<div class="tabella scorri"><table>
<thead><tr><th>File</th><th>Cliente</th><th>Caricato</th><th>Stato</th><th></th></tr></thead><tbody>
{% for doc in elenco %}<tr>
  <td>{% if doc.link_drive %}<a href="{{ doc.link_drive }}" target="_blank">{{ doc.nome_file }}</a>
      {% else %}{{ doc.nome_file }}{% endif %}</td>
  <td><a href="/crm/clienti/{{ doc.cliente_id }}">{{ doc.cliente.ragione_sociale }}</a></td>
  <td>{{ data_it(doc.creato_il.date()) }} · {{ doc.caricato_da }}</td>
  <td>{% if doc.stato == 'assegnato' %}<span class="pill ok">→ {{ doc.pratica.nome_bando if doc.pratica else 'assegnato' }}</span>
      {% else %}<span class="pill">da smistare</span>{% endif %}</td>
  <td class="num">
    {% if doc.stato != 'assegnato' and doc.cliente.pratiche %}
    <form method="post" action="/crm/documenti/{{ doc.id }}/assegna" style="display:inline-flex;gap:6px">
      <select name="pratica_id" required><option value="">Assegna a…</option>
        {% for p in doc.cliente.pratiche %}<option value="{{ p.id }}">{{ p.nome_bando }}</option>{% endfor %}</select>
      <button class="btn chiaro" type="submit">Ok</button>
    </form>
    {% elif doc.stato != 'assegnato' %}<span class="nota">Il cliente non ha ancora pratiche</span>{% endif %}
  </td>
</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="vuoto">{{ 'Niente da smistare al momento.' if solo_da_smistare else 'Nessun documento caricato finora.' }}</div>{% endif %}
{% endblock %}"""

T_BANDI = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>Bandi</h1>
<p class="sottotitolo">{{ elenco|length }} nel repository · da qui nascono le pratiche già precompilate</p></div>
<div><a class="btn" href="/crm/bandi/nuovo">Nuovo bando</a></div></div>

<form class="filtri" method="get">
  <input name="q" value="{{ q or '' }}" placeholder="Cerca nome o ente" style="min-width:260px">
  <button class="btn" type="submit">Filtra</button>
</form>

{% if not configurato %}
<div class="vuoto" style="margin-bottom:16px">L'estrazione da PDF e la generazione di guide/approfondimenti richiedono
ANTHROPIC_API_KEY, non ancora configurata. Puoi comunque creare bandi a mano.</div>
{% endif %}

{% if elenco %}
<div class="tabella scorri"><table>
<thead><tr><th>Nome</th><th>Ente</th><th>Tipologia</th><th>Contributo max</th><th>Scadenza</th><th></th></tr></thead><tbody>
{% for b in elenco %}<tr>
  <td><a href="/crm/bandi/{{ b.id }}">{{ b.nome }}</a></td>
  <td>{{ b.ente or '—' }}</td>
  <td>{{ b.tipologia or '—' }}</td>
  <td>{{ b.importo_max_testo or euro(b.contributo_max) }}</td>
  <td>{{ data_it(b.data_scadenza) }}</td>
  <td class="num"><a class="btn chiaro" href="/crm/pratiche/nuova?bando={{ b.id }}">Crea pratica</a></td>
</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="vuoto">Nessun bando nel repository ancora.</div>{% endif %}
{% endblock %}"""

T_BANDO_FORM = """{% extends "base" %}{% block contenuto %}
<h1>{{ 'Modifica ' ~ b.nome if b.id else 'Nuovo bando' }}</h1>

{% if not b.id %}
<div class="riquadro" style="margin-bottom:20px"><div class="corpo">
  <h2 style="margin-top:0">Carica una scheda PDF</h2>
  {% if configurato %}
  <p style="color:#6b7b8c">Un'IA legge la scheda e prova a compilare i campi da sola — controlli e correggi prima di salvare.</p>
  <form method="post" action="/crm/bandi/nuovo-da-pdf" enctype="multipart/form-data" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <input type="file" name="file" accept=".pdf" required>
    <button class="btn ambra" type="submit">Estrai dal PDF</button>
  </form>
  {% else %}
  <p style="color:#6b7b8c">Richiede ANTHROPIC_API_KEY, non ancora configurata. Compila i campi a mano qui sotto.</p>
  {% endif %}
</div></div>
<p class="sottotitolo">— oppure compila a mano —</p>
{% endif %}

<form method="post">
<fieldset><legend>Dati principali</legend><div class="griglia g3">
  <label><span class="etichetta">Nome *</span><input name="nome" required value="{{ b.nome or '' }}"></label>
  <label><span class="etichetta">Ente</span><input name="ente" value="{{ b.ente or '' }}"></label>
  <label><span class="etichetta">Tipologia</span><select name="tipologia"><option value="">—</option>
    {% for t in tipologie %}<option {{ 'selected' if b.tipologia==t }}>{{ t }}</option>{% endfor %}</select></label>
  <label><span class="etichetta">Dotazione €</span><input name="dotazione" inputmode="decimal" value="{{ b.dotazione or '' }}"></label>
  <label><span class="etichetta">% contributo</span><input name="perc_contributo" inputmode="decimal" value="{{ b.perc_contributo or '' }}"></label>
  <label><span class="etichetta">Contributo max €</span><input name="contributo_max" inputmode="decimal" value="{{ b.contributo_max or '' }}"></label>
  <label><span class="etichetta">Importo max (testo libero)</span><input name="importo_max_testo" value="{{ b.importo_max_testo or '' }}" placeholder="Es. 30k linea A / 40k linea B"></label>
  <label><span class="etichetta">Apertura</span><input name="data_apertura" type="date" value="{{ b.data_apertura.isoformat() if b.data_apertura else '' }}"></label>
  <label><span class="etichetta">Scadenza</span><input name="data_scadenza" type="date" value="{{ b.data_scadenza.isoformat() if b.data_scadenza else '' }}"></label>
</div></fieldset>
<fieldset><legend>Scheda discorsiva</legend>
  <label><span class="etichetta">Chi può partecipare</span><textarea name="chi_puo_partecipare">{{ b.chi_puo_partecipare or '' }}</textarea></label>
  <label><span class="etichetta">Cosa è finanziabile</span><textarea name="cosa_finanziabile">{{ b.cosa_finanziabile or '' }}</textarea></label>
  <label><span class="etichetta">Spese non ammissibili</span><textarea name="spese_non_ammissibili">{{ b.spese_non_ammissibili or '' }}</textarea></label>
  <label><span class="etichetta">Criteri di valutazione</span><textarea name="criteri">{{ b.criteri or '' }}</textarea></label>
  <label><span class="etichetta">Fasi e tempi</span><textarea name="fasi_tempi">{{ b.fasi_tempi or '' }}</textarea></label>
  <label><span class="etichetta">Come presentare</span><textarea name="come_presentare">{{ b.come_presentare or '' }}</textarea></label>
  <label><span class="etichetta">Perché è interessante</span><textarea name="perche_interessante">{{ b.perche_interessante or '' }}</textarea></label>
  <label><span class="etichetta">Criticità</span><textarea name="criticita">{{ b.criticita or '' }}</textarea></label>
</fieldset>
<button class="btn" type="submit">Salva</button>
</form>
{% endblock %}"""

T_BANDO = """{% extends "base" %}{% block contenuto %}
<div class="testa"><div><h1>{{ b.nome }}</h1>
<p class="sottotitolo">{{ b.ente or '—' }}{% if b.tipologia %} · {{ b.tipologia }}{% endif %}</p></div>
<div><a class="btn chiaro" href="/crm/bandi/{{ b.id }}/modifica">Modifica</a>
<a class="btn" href="/crm/pratiche/nuova?bando={{ b.id }}">Crea pratica per un cliente</a></div></div>

<div class="griglia g4" style="margin-bottom:20px">
  <div class="kpi"><div class="etichetta">Dotazione</div><div class="valore" style="font-size:19px">{{ euro(b.dotazione) }}</div></div>
  <div class="kpi"><div class="etichetta">% contributo</div><div class="valore" style="font-size:19px">{{ (b.perc_contributo|string ~ ' %') if b.perc_contributo else '—' }}</div></div>
  <div class="kpi"><div class="etichetta">Contributo max</div><div class="valore" style="font-size:19px">{{ b.importo_max_testo or euro(b.contributo_max) }}</div></div>
  <div class="kpi"><div class="etichetta">Scadenza</div><div class="valore" style="font-size:19px">{{ data_it(b.data_scadenza) }}</div></div>
</div>

<div class="griglia g2">
<div>
  <h2 style="margin-top:0">Chi può partecipare</h2><p style="white-space:pre-wrap">{{ b.chi_puo_partecipare or '—' }}</p>
  <h2>Cosa è finanziabile</h2><p style="white-space:pre-wrap">{{ b.cosa_finanziabile or '—' }}</p>
  <h2>Spese non ammissibili</h2><p style="white-space:pre-wrap">{{ b.spese_non_ammissibili or '—' }}</p>
  <h2>Criteri di valutazione</h2><p style="white-space:pre-wrap">{{ b.criteri or '—' }}</p>
</div>
<div>
  <h2 style="margin-top:0">Fasi e tempi</h2><p style="white-space:pre-wrap">{{ b.fasi_tempi or '—' }}</p>
  <h2>Come presentare</h2><p style="white-space:pre-wrap">{{ b.come_presentare or '—' }}</p>
  <h2>Perché è interessante</h2><p style="white-space:pre-wrap">{{ b.perche_interessante or '—' }}</p>
  <h2>Criticità</h2><p style="white-space:pre-wrap">{{ b.criticita or '—' }}</p>
</div>
</div>

<h2>Guida alla compilazione</h2>
<div class="riquadro"><div class="corpo">
{% if b.guida_compilazione %}
  <p style="white-space:pre-wrap">{{ b.guida_compilazione }}</p>
  <a class="btn chiaro" href="/crm/bandi/{{ b.id }}/guida.txt">Scarica la guida</a>
  <form method="post" action="/crm/bandi/{{ b.id }}/guida" style="display:inline"><button class="btn chiaro" type="submit">Rigenera</button></form>
{% elif configurato %}
  <p style="margin-top:0;color:#6b7b8c">Non ancora generata.</p>
  <form method="post" action="/crm/bandi/{{ b.id }}/guida"><button class="btn ambra" type="submit">Genera guida di compilazione</button></form>
{% else %}
  <p style="margin-top:0;color:#6b7b8c">Richiede ANTHROPIC_API_KEY, non ancora configurata.</p>
{% endif %}
</div></div>

<h2>Approfondimento</h2>
<div class="riquadro"><div class="corpo">
{% if b.approfondimento %}
  <p style="white-space:pre-wrap">{{ b.approfondimento }}</p>
  <form method="post" action="/crm/bandi/{{ b.id }}/approfondisci" style="display:inline"><button class="btn chiaro" type="submit">Rigenera</button></form>
{% elif configurato %}
  <p style="margin-top:0;color:#6b7b8c">Non ancora generato. Compare nella scheda di ogni pratica nata da questo bando.</p>
  <form method="post" action="/crm/bandi/{{ b.id }}/approfondisci"><button class="btn ambra" type="submit">Genera approfondimento</button></form>
{% else %}
  <p style="margin-top:0;color:#6b7b8c">Richiede ANTHROPIC_API_KEY, non ancora configurata.</p>
{% endif %}
</div></div>
{% endblock %}"""

T_IMPOSTAZIONI = """{% extends "base" %}{% block contenuto %}
<h1>Impostazioni</h1><p class="sottotitolo">Solo gli amministratori vedono questa pagina.</p>

<h2>Utenti</h2>
<div class="tabella scorri"><table>
<thead><tr><th>Nome</th><th>Email</th><th>Ruolo</th><th>Stato</th><th></th></tr></thead><tbody>
{% for u in utenti %}<tr>
  <td>{{ u.nome }}</td><td>{{ u.email }}</td><td>{{ u.ruolo }}</td>
  <td>{{ 'attivo' if u.attivo else 'disattivato' }}</td>
  <td class="num">{% if u.id != utente.id %}
    <form method="post" action="/crm/utenti/{{ u.id }}/stato" style="display:inline">
    <button class="btn chiaro" type="submit">{{ 'Disattiva' if u.attivo else 'Riattiva' }}</button></form>{% endif %}</td>
</tr>{% endfor %}</tbody></table></div>

<h2>Nuovo utente</h2>
<form method="post" action="/crm/utenti/nuovo" class="riquadro"><div class="corpo"><div class="griglia g4">
  <label><span class="etichetta">Nome e cognome</span><input name="nome" required></label>
  <label><span class="etichetta">Email</span><input name="email" type="email" required></label>
  <label><span class="etichetta">Password</span>
  <div class="campo-pw"><input id="pw-nuovo" name="password" type="password" required minlength="8">
  <button type="button" class="occhio" onclick="mostraPw('pw-nuovo')" aria-label="Mostra password">occhio</button></div></label>
  <label><span class="etichetta">Ruolo</span><select name="ruolo">
    {% for r in ruoli %}<option>{{ r }}</option>{% endfor %}</select></label>
</div><button class="btn" type="submit">Crea utente</button></div></form>

<h2>Conti bancari Energelia</h2>
<p class="sottotitolo" style="margin-top:-8px">Su questi conti scegliete di volta in volta, per ogni pratica, su quale il cliente deve pagarvi.</p>
<div class="tabella scorri"><table>
<thead><tr><th>Nome</th><th>IBAN</th><th>Banca</th><th>Stato</th><th></th></tr></thead><tbody>
{% for conto in conti_tutti %}<tr>
  <td>{{ conto.nome }}</td><td>{{ conto.iban or '—' }}</td><td>{{ conto.banca or '—' }}</td>
  <td>{{ 'attivo' if conto.attivo else 'disattivato' }}</td>
  <td class="num">
    <form method="post" action="/crm/conti/{{ conto.id }}/stato" style="display:inline">
    <button class="btn chiaro" type="submit">{{ 'Disattiva' if conto.attivo else 'Riattiva' }}</button></form></td>
</tr>{% endfor %}</tbody></table></div>

<h2>Nuovo conto</h2>
<form method="post" action="/crm/conti/nuovo" class="riquadro"><div class="corpo"><div class="griglia g3">
  <label><span class="etichetta">Nome identificativo</span><input name="nome" required placeholder="Es. Energelia — Ordinario"></label>
  <label><span class="etichetta">IBAN</span><input name="iban"></label>
  <label><span class="etichetta">Banca</span><input name="banca"></label>
</div><button class="btn" type="submit">Aggiungi conto</button></div></form>

<h2>Cambia la tua password</h2>
<form method="post" action="/crm/password" class="riquadro"><div class="corpo"><div class="griglia g2">
  <label><span class="etichetta">Password attuale</span>
  <div class="campo-pw"><input id="pw-attuale" name="attuale" type="password" required>
  <button type="button" class="occhio" onclick="mostraPw('pw-attuale')" aria-label="Mostra password">occhio</button></div></label>
  <label><span class="etichetta">Nuova password</span>
  <div class="campo-pw"><input id="pw-nuova" name="nuova" type="password" required minlength="8">
  <button type="button" class="occhio" onclick="mostraPw('pw-nuova')" aria-label="Mostra password">occhio</button></div></label>
</div><button class="btn" type="submit">Aggiorna password</button></div></form>

<h2>Importa dal foglio Excel</h2>
<form method="post" action="/crm/importa" enctype="multipart/form-data" class="riquadro"><div class="corpo">
  <p style="margin-top:0;color:#6b7b8c">Carica <em>Energelia_CRM_Clienti.xlsx</em>. Legge i fogli
  <strong>Clienti</strong> e <strong>Pratiche</strong> con le intestazioni in riga 4. I clienti già
  presenti (stessa ragione sociale) vengono saltati, non duplicati.</p>
  <input type="file" name="file" accept=".xlsx" required>
  <button class="btn ambra" type="submit" style="margin-left:8px">Importa</button>
</div></form>

<h2>Importa lead da scraping</h2>
<form method="post" action="/crm/lead/importa" enctype="multipart/form-data" class="riquadro"><div class="corpo">
  <p style="margin-top:0;color:#6b7b8c">Carica direttamente l'output di uno degli scraper (Nome, Tipo, Indirizzo,
  Tel. fisso, Cellulare, Sito web, Email...). Riconosce le intestazioni da sé, salta le righe separatore
  di sezione, e non duplica un nome già importato con lo stesso indirizzo.</p>
  <input name="fonte" placeholder="Etichetta di questo giro, es. Valle d'Aosta - Ristoranti" style="min-width:280px">
  <input type="file" name="file" accept=".xlsx" required>
  <button class="btn ambra" type="submit" style="margin-left:8px">Importa lead</button>
</div></form>

<h2>Esporta</h2>
<p><a class="btn chiaro" href="/crm/esporta.xlsx">Scarica tutto in Excel</a>
<a class="btn chiaro" href="/crm/esporta.csv">Clienti in CSV</a></p>
{% endblock %}"""

env = Environment(loader=DictLoader({
    "base": BASE, "login": T_LOGIN, "dashboard": T_DASHBOARD, "clienti": T_CLIENTI,
    "cliente_form": T_CLIENTE_FORM, "cliente": T_CLIENTE, "pratiche": T_PRATICHE,
    "pratica_form": T_PRATICA_FORM, "pratica": T_PRATICA, "attivita": T_ATTIVITA,
    "impostazioni": T_IMPOSTAZIONI, "lead": T_LEAD, "carica_pubblico": T_CARICA_PUBBLICO,
    "richiesta_pubblica": T_RICHIESTA_PUBBLICA,
    "documenti": T_DOCUMENTI, "bandi": T_BANDI, "bando_form": T_BANDO_FORM, "bando": T_BANDO,
}), autoescape=select_autoescape(["html"]))


def classe_fase(fase):
    if fase in FASI_VINTE:
        return "fase-ok"
    if fase == "Respinta":
        return "fase-ko"
    return "fase"


env.globals.update(euro=euro, data_it=data_it, classe_fase=classe_fase,
                   fasi=FASI, canali=CANALI, tipologie=TIPOLOGIE, priorita_lista=PRIORITA,
                   dimensioni=DIMENSIONI, tipi=TIPI_ATTIVITA, ruoli=RUOLI,
                   oggi_iso=lambda: dt.date.today().isoformat())


def rendi(nome, **ctx):
    ctx.setdefault("titolo", "CRM")
    ctx["utente"] = utente_corrente()
    ctx["messaggi"] = session.pop("crm_messaggi", [])
    ctx["oggi_iso"] = dt.date.today().isoformat()
    ctx["request"] = request
    return env.get_template(nome).render(**ctx)


def avvisa(testo, tipo="ok"):
    messaggi = session.get("crm_messaggi", [])
    messaggi.append([tipo, testo])
    session["crm_messaggi"] = messaggi


# --------------------------------------------------------------------------
# BLUEPRINT
# --------------------------------------------------------------------------

crm = Blueprint("crm", __name__, url_prefix="/crm")

LIBERE = {"crm.accedi", "crm.pagina_caricamento", "crm.carica_file_pubblico",
          "crm.pagina_richiesta", "crm.invia_richiesta"}


def utente_corrente():
    uid = session.get("crm_uid")
    if not uid:
        return None
    u = SessionLocale.get(Utente, uid)
    return u if (u and u.attivo) else None


@crm.before_request
def richiedi_accesso():
    if request.endpoint in LIBERE:
        return None
    if utente_corrente() is None:
        session.pop("crm_uid", None)
        return env.get_template("login").render(errore=None)
    return None


@crm.post("/accedi")
def accedi():
    email = (request.form.get("email") or "").strip().lower()
    u = SessionLocale.query(Utente).filter(func.lower(Utente.email) == email).first()
    if not u or not u.attivo or not verifica_pw(request.form.get("password") or "", u.password_hash):
        return env.get_template("login").render(errore="Email o password non corrette."), 401
    session["crm_uid"] = u.id
    session.permanent = True
    return redirect("/crm/")


@crm.get("/esci")
def esci():
    session.pop("crm_uid", None)
    session.pop("crm_messaggi", None)
    return redirect("/crm/")


# --------------------------------------------------- caricamento pubblico

@crm.get("/carica/<token>")
def pagina_caricamento(token):
    cliente = SessionLocale.query(Cliente).filter_by(token_caricamento=token).first()
    if not cliente:
        abort(404)
    return env.get_template("carica_pubblico").render(cliente=cliente, configurato=drive_configurato())


@crm.post("/carica/<token>")
def carica_file_pubblico(token):
    cliente = SessionLocale.query(Cliente).filter_by(token_caricamento=token).first()
    if not cliente:
        abort(404)
    if not drive_configurato():
        return env.get_template("carica_pubblico").render(cliente=cliente, configurato=False)

    ricevuti = 0
    for f in request.files.getlist("file"):
        if not f or not f.filename:
            continue
        try:
            cartella_id = _drive_cartella_cliente(cliente)
            risultato = _drive_carica_file(cartella_id, f.filename, f.read(), f.mimetype)
            SessionLocale.add(Documento(
                cliente_id=cliente.id, nome_file=f.filename,
                google_file_id=risultato.get("id"), link_drive=risultato.get("webViewLink"),
                dimensione_byte=int(risultato.get("size") or 0),
                caricato_da="cliente", stato="da_smistare"))
            ricevuti += 1
        except Exception as errore:
            print(f"[crm] Errore caricamento file '{f.filename}' da {cliente.codice}: {errore}")
    SessionLocale.commit()
    return env.get_template("carica_pubblico").render(cliente=cliente, configurato=True, fatto=ricevuti)


@crm.get("/richiesta/<token>")
def pagina_richiesta(token):
    p = SessionLocale.query(Pratica).filter_by(token_caricamento=token).first()
    if not p:
        abort(404)
    return env.get_template("richiesta_pubblica").render(p=p, configurato=drive_configurato())


@crm.post("/richiesta/<token>")
def invia_richiesta(token):
    p = SessionLocale.query(Pratica).filter_by(token_caricamento=token).first()
    if not p:
        abort(404)
    if not drive_configurato():
        return env.get_template("richiesta_pubblica").render(p=p, configurato=False)

    for v in p.voci_richiesta:
        if v.tipo_risposta == "testo":
            valore = s(request.form.get(f"voce_{v.id}"))
            if valore:
                v.valore_testo = valore
                v.compilata = True
        else:
            file_caricati = [f for f in request.files.getlist(f"voce_{v.id}") if f and f.filename]
            for f in file_caricati:
                try:
                    cartella_id = _drive_cartella_cliente(p.cliente)
                    risultato = _drive_carica_file(cartella_id, f.filename, f.read(), f.mimetype)
                    doc = Documento(
                        cliente_id=p.cliente_id, pratica_id=p.id, nome_file=f.filename,
                        google_file_id=risultato.get("id"), link_drive=risultato.get("webViewLink"),
                        dimensione_byte=int(risultato.get("size") or 0),
                        caricato_da="cliente", stato="assegnato")
                    SessionLocale.add(doc)
                    SessionLocale.flush()
                    if not v.documento_id:      # il link della voce punta al primo file; gli altri
                        v.documento_id = doc.id  # restano comunque visibili tra i documenti della pratica
                    v.compilata = True
                except Exception as errore:
                    print(f"[crm] Errore caricamento voce '{v.etichetta}' pratica {p.codice}: {errore}")

    for f in request.files.getlist("extra"):
        if not f or not f.filename:
            continue
        try:
            cartella_id = _drive_cartella_cliente(p.cliente)
            risultato = _drive_carica_file(cartella_id, f.filename, f.read(), f.mimetype)
            SessionLocale.add(Documento(
                cliente_id=p.cliente_id, pratica_id=p.id, nome_file=f.filename,
                google_file_id=risultato.get("id"), link_drive=risultato.get("webViewLink"),
                dimensione_byte=int(risultato.get("size") or 0),
                caricato_da="cliente", stato="assegnato"))
        except Exception as errore:
            print(f"[crm] Errore caricamento extra pratica {p.codice}: {errore}")

    SessionLocale.commit()
    return env.get_template("richiesta_pubblica").render(p=p, configurato=True, fatto=True)


# ---------------------------------------------------------------- dashboard

@crm.get("/")
def dashboard():
    pratiche = SessionLocale.query(Pratica).all()
    clienti = SessionLocale.query(Cliente).all()

    def somma(campo):
        return sum((getattr(p, campo) or Decimal(0)) for p in pratiche)

    k = {
        "clienti": len(clienti),
        "clienti_con_pratica": sum(1 for c in clienti if c.pratiche),
        "pratiche": len(pratiche),
        "pratiche_aperte": sum(1 for p in pratiche if p.fase in FASI_APERTE),
        "richiesto": somma("importo_richiesto"),
        "concesso": somma("importo_concesso"),
        "ammesse": sum(1 for p in pratiche if p.fase in FASI_VINTE),
        "corrispettivi": somma("corrispettivo"),
        "success_fee": sum(p.success_fee_maturata for p in pratiche),
        "incassato": somma("incassato"),
    }

    conteggi = {f: 0 for f in FASI}
    for p in pratiche:
        if p.fase in conteggi:
            conteggi[p.fase] += 1
    pipeline = [(f, FASI_BREVI.get(f, f), conteggi[f]) for f in FASI]
    massimo = max([c for *_, c in pipeline] + [1])

    limite = dt.date.today() + dt.timedelta(days=60)
    scadenze = sorted([p for p in pratiche
                       if p.data_scadenza and p.fase in FASI_PRE_INVIO and p.data_scadenza <= limite],
                      key=lambda p: p.data_scadenza)[:10]

    fermi = []
    for c in clienti:
        ultimo = c.ultimo_contatto or c.data_primo_contatto
        giorni = (dt.date.today() - ultimo).days if ultimo else None
        if giorni is None or giorni >= 30:
            fermi.append((c, giorni))
    fermi.sort(key=lambda x: (-1 if x[1] is None else -x[1]))

    attivita = SessionLocale.query(Attivita).order_by(
        Attivita.data.desc(), Attivita.id.desc()).limit(8).all()

    return rendi("dashboard", titolo="Dashboard", pagina="dashboard", k=k,
                 pipeline=pipeline, massimo=massimo, scadenze=scadenze, fermi=fermi[:8],
                 attivita=attivita, oggi=dt.date.today().strftime("%d/%m/%Y"))


# ------------------------------------------------------------------ clienti

@crm.get("/clienti")
def lista_clienti():
    q = request.args.get("q", "")
    canale = request.args.get("canale", "")
    consulente = request.args.get("consulente", "")
    query = SessionLocale.query(Cliente)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Cliente.ragione_sociale.ilike(like), Cliente.piva.ilike(like),
                                 Cliente.referente.ilike(like), Cliente.citta.ilike(like),
                                 Cliente.codice.ilike(like)))
    if canale:
        query = query.filter(Cliente.canale == canale)
    if consulente:
        query = query.filter(Cliente.consulente_id == int(consulente))
    return rendi("clienti", titolo="Clienti", pagina="clienti",
                 elenco=query.order_by(Cliente.ragione_sociale).all(),
                 totale=SessionLocale.query(func.count(Cliente.id)).scalar(),
                 q=q, canale=canale, consulente=consulente,
                 consulenti=SessionLocale.query(Utente).filter(
                     Utente.attivo.is_(True)).order_by(Utente.nome).all())


def _leggi_cliente(form, cliente):
    cliente.ragione_sociale = s(form.get("ragione_sociale"))
    cliente.piva = s(form.get("piva"))
    cliente.ateco = s(form.get("ateco"))
    cliente.dimensione = s(form.get("dimensione"))
    cliente.citta = s(form.get("citta"))
    cliente.provincia = (s(form.get("provincia")) or "").upper() or None
    cliente.regione = s(form.get("regione"))
    cliente.referente = s(form.get("referente"))
    cliente.ruolo_referente = s(form.get("ruolo_referente"))
    cliente.telefono = s(form.get("telefono"))
    cliente.email = s(form.get("email"))
    cliente.pec = s(form.get("pec"))
    cliente.codice_fiscale = s(form.get("codice_fiscale"))
    cliente.codice_sdi = s(form.get("codice_sdi"))
    cliente.intestatario_conto = s(form.get("intestatario_conto"))
    cliente.iban = (s(form.get("iban")) or "").upper().replace(" ", "") or None
    cliente.email_fatturazione = s(form.get("email_fatturazione"))
    cliente.titolari_effettivi = s(form.get("titolari_effettivi"))
    cid = s(form.get("consulente_id"))
    cliente.consulente_id = int(cid) if cid else None
    cliente.canale = s(form.get("canale"))
    cliente.data_primo_contatto = d(form.get("data_primo_contatto"))
    cliente.prossima_azione = s(form.get("prossima_azione"))
    cliente.note = s(form.get("note"))


@crm.get("/clienti/nuovo")
def nuovo_cliente_form():
    cliente = Cliente()
    lead = None
    lead_id = request.args.get("lead", "")
    if lead_id:
        lead = SessionLocale.get(Lead, int(lead_id))
        if lead:
            cliente.ragione_sociale = lead.nome
            cliente.telefono = lead.cellulare or lead.telefono
            cliente.email = lead.email
            cliente.pec = lead.pec
            cliente.canale = "Scrapping"
            note = []
            if lead.sito: note.append(f"Sito: {lead.sito}")
            if lead.indirizzo: note.append(f"Indirizzo: {lead.indirizzo}")
            if lead.fonte: note.append(f"Da lead — {lead.fonte}")
            cliente.note = "\n".join(note) or None
    return rendi("cliente_form", titolo="Nuovo cliente", pagina="clienti", cliente=cliente, lead=lead,
                 anthropic_ok=anthropic_configurato(),
                 consulenti=SessionLocale.query(Utente).order_by(Utente.nome).all())


@crm.post("/clienti/nuovo-da-visura")
def nuovo_cliente_da_visura():
    lead = None
    lead_id = s(request.form.get("lead_id"))
    if lead_id:
        lead = SessionLocale.get(Lead, int(lead_id))

    if not anthropic_configurato():
        avvisa("L'estrazione dalla visura non è configurata: manca ANTHROPIC_API_KEY.", "ko")
        return redirect(f"/crm/clienti/nuovo?lead={lead_id}" if lead_id else "/crm/clienti/nuovo")

    caricato = request.files.get("file")
    if not caricato or not caricato.filename:
        avvisa("Scegli un PDF da caricare.", "ko")
        return redirect(f"/crm/clienti/nuovo?lead={lead_id}" if lead_id else "/crm/clienti/nuovo")

    cliente = Cliente()
    # Se arriva da un lead, parto dagli stessi dati che avresti visto senza la visura.
    if lead:
        cliente.ragione_sociale = lead.nome
        cliente.telefono = lead.cellulare or lead.telefono
        cliente.email = lead.email
        cliente.pec = lead.pec
        cliente.canale = "Scrapping"

    try:
        testo = _estrai_testo_pdf(caricato.read())
        dati = estrai_visura_da_testo(testo)
    except Exception as errore:
        avvisa(f"Estrazione non riuscita: {errore}", "ko")
        return rendi("cliente_form", titolo="Nuovo cliente", pagina="clienti", cliente=cliente, lead=lead,
                     anthropic_ok=anthropic_configurato(),
                     consulenti=SessionLocale.query(Utente).order_by(Utente.nome).all())

    for campo in ("ragione_sociale", "piva", "codice_fiscale", "ateco", "citta", "provincia",
                  "regione", "pec", "referente", "ruolo_referente", "titolari_effettivi"):
        valore = (dati.get(campo) or "").strip()
        if valore:
            setattr(cliente, campo, valore)

    avvisa("Dati presi dalla visura: controlla i campi prima di salvare.")
    return rendi("cliente_form", titolo="Nuovo cliente", pagina="clienti", cliente=cliente, lead=lead,
                 anthropic_ok=anthropic_configurato(), da_visura=True,
                 consulenti=SessionLocale.query(Utente).order_by(Utente.nome).all())


@crm.post("/clienti/nuovo")
def nuovo_cliente():
    if not s(request.form.get("ragione_sociale")):
        avvisa("Serve almeno la ragione sociale.", "ko")
        return redirect("/crm/clienti/nuovo")
    c = Cliente(codice=prossimo_codice(SessionLocale, Cliente, "CL"))
    _leggi_cliente(request.form, c)
    SessionLocale.add(c)
    SessionLocale.flush()

    lead_id = s(request.form.get("lead_id"))
    if lead_id:
        lead = SessionLocale.get(Lead, int(lead_id))
        if lead:
            lead.stato = "convertito"
            lead.cliente_id = c.id

    SessionLocale.commit()
    avvisa(f"Cliente {c.ragione_sociale} salvato con codice {c.codice}.")
    return redirect(f"/crm/clienti/{c.id}")


@crm.get("/clienti/<int:cid>")
def scheda_cliente(cid):
    c = SessionLocale.get(Cliente, cid)
    if not c:
        abort(404)
    if not c.token_caricamento:
        c.token_caricamento = secrets.token_urlsafe(24)
        SessionLocale.commit()
    documenti = (SessionLocale.query(Documento).filter_by(cliente_id=cid)
                 .order_by(Documento.creato_il.desc()).all())
    return rendi("cliente", titolo=c.ragione_sociale, pagina="clienti", c=c, documenti=documenti)


@crm.post("/clienti/<int:cid>/invia-link")
def invia_link_cliente(cid):
    c = SessionLocale.get(Cliente, cid)
    destinatario = s(request.form.get("destinatario"))
    if not c or not destinatario:
        avvisa("Serve un indirizzo email valido.", "ko")
        return redirect(f"/crm/clienti/{cid}")
    io_stesso = utente_corrente()
    link = f"{request.url_root.rstrip('/')}/crm/carica/{c.token_caricamento}"
    corpo = (
        f"Buongiorno,\n\n"
        f"può caricare i documenti che ci servono per {c.ragione_sociale} da questo link, "
        f"anche in più volte e senza un ordine preciso:\n\n{link}\n\n"
        f"Grazie,\n{io_stesso.nome}\nEnergelia S.r.l."
    )
    ok, errore = invia_email(destinatario, f"Documenti per {c.ragione_sociale}", corpo,
                             rispondi_a=io_stesso.email, nome_mittente="Energelia")
    avvisa("Email inviata." if ok else f"Invio non riuscito: {errore}", "ok" if ok else "ko")
    return redirect(f"/crm/clienti/{cid}")


@crm.get("/clienti/<int:cid>/modifica")
def modifica_cliente_form(cid):
    c = SessionLocale.get(Cliente, cid)
    if not c:
        abort(404)
    return rendi("cliente_form", titolo="Modifica cliente", pagina="clienti", cliente=c,
                 anthropic_ok=anthropic_configurato(),
                 consulenti=SessionLocale.query(Utente).order_by(Utente.nome).all())


@crm.post("/clienti/<int:cid>/cerca-online")
def cerca_online_cliente(cid):
    c = SessionLocale.get(Cliente, cid)
    if not c:
        abort(404)
    if not anthropic_configurato():
        avvisa("La ricerca online non è configurata: manca ANTHROPIC_API_KEY.", "ko")
        return redirect(f"/crm/clienti/{cid}/modifica")
    try:
        trovato = cerca_cliente_online(c)
    except Exception as errore:
        avvisa(f"Ricerca non riuscita: {errore}", "ko")
        return redirect(f"/crm/clienti/{cid}/modifica")

    # Non committo: aggiorno l'oggetto solo in memoria per farlo vedere nel
    # modulo di modifica. Se l'utente non salva, la sessione si chiude a fine
    # richiesta e questi valori non toccati non finiscono mai sul database.
    trovati = []
    mappa_campi = {
        "telefono": "telefono", "email": "email", "pec": "pec",
        "citta": "citta", "provincia": "provincia",
    }
    for chiave_ricerca, campo_cliente in mappa_campi.items():
        valore = (trovato.get(chiave_ricerca) or "").strip()
        if valore and not getattr(c, campo_cliente):   # non sovrascrivo dati già presenti
            setattr(c, campo_cliente, valore)
            trovati.append(campo_cliente)
    sito_trovato = (trovato.get("sito") or "").strip()
    indirizzo_trovato = (trovato.get("indirizzo") or "").strip()
    note_ricerca = []
    if sito_trovato: note_ricerca.append(f"Sito trovato online: {sito_trovato}")
    if indirizzo_trovato: note_ricerca.append(f"Indirizzo trovato online: {indirizzo_trovato}")
    if note_ricerca:
        c.note = ((c.note + "\n") if c.note else "") + "\n".join(note_ricerca)

    if trovati or note_ricerca:
        avvisa(f"Trovati online: {', '.join(trovati) if trovati else 'sito/indirizzo in nota'}. Controlla e salva.")
    else:
        avvisa("Nessuna informazione nuova trovata online.", "ko")

    return rendi("cliente_form", titolo="Modifica cliente", pagina="clienti", cliente=c,
                 anthropic_ok=anthropic_configurato(), da_ricerca=bool(trovati or note_ricerca),
                 consulenti=SessionLocale.query(Utente).order_by(Utente.nome).all())


@crm.post("/clienti/<int:cid>/modifica")
def modifica_cliente(cid):
    c = SessionLocale.get(Cliente, cid)
    if not c:
        abort(404)
    _leggi_cliente(request.form, c)
    SessionLocale.commit()
    avvisa("Modifiche salvate.")
    return redirect(f"/crm/clienti/{c.id}")


# ----------------------------------------------------------------- pratiche

@crm.get("/pratiche")
def lista_pratiche():
    q = request.args.get("q", "")
    fase = request.args.get("fase", "")
    priorita = request.args.get("priorita", "")
    aperte = request.args.get("aperte", "")
    query = SessionLocale.query(Pratica).join(Cliente)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Pratica.nome_bando.ilike(like), Pratica.ente.ilike(like),
                                 Pratica.codice.ilike(like), Cliente.ragione_sociale.ilike(like)))
    if fase:
        query = query.filter(Pratica.fase == fase)
    if priorita:
        query = query.filter(Pratica.priorita == priorita)
    if aperte:
        query = query.filter(Pratica.fase.in_(FASI_APERTE))
    elenco = query.order_by(Pratica.data_scadenza.is_(None), Pratica.data_scadenza).all()
    return rendi("pratiche", titolo="Pratiche", pagina="pratiche", elenco=elenco,
                 totale=SessionLocale.query(func.count(Pratica.id)).scalar(),
                 q=q, fase=fase, priorita=priorita, aperte=bool(aperte),
                 tot_richiesto=sum((p.importo_richiesto or Decimal(0)) for p in elenco),
                 tot_concesso=sum((p.importo_concesso or Decimal(0)) for p in elenco))


def _leggi_pratica(form, p):
    p.cliente_id = int(form.get("cliente_id"))
    bando = s(form.get("bando_id"))
    if bando:
        p.bando_id = int(bando)
    p.nome_bando = s(form.get("nome_bando"))
    p.ente = s(form.get("ente"))
    p.tipologia = s(form.get("tipologia"))
    p.perc_contributo = n(form.get("perc_contributo"))
    p.importo_max = s(form.get("importo_max"))
    p.data_apertura = d(form.get("data_apertura"))
    p.data_scadenza = d(form.get("data_scadenza"))
    p.fase = s(form.get("fase")) or "Analisi fattibilità"
    p.data_presentazione = d(form.get("data_presentazione"))
    p.data_esito = d(form.get("data_esito"))
    p.importo_richiesto = n(form.get("importo_richiesto"))
    p.importo_concesso = n(form.get("importo_concesso"))
    p.scadenza_rendicontazione = d(form.get("scadenza_rendicontazione"))
    p.documenti_mancanti = s(form.get("documenti_mancanti"))
    p.prossimo_step = s(form.get("prossimo_step"))
    p.priorita = s(form.get("priorita")) or "Media"
    p.corrispettivo = n(form.get("corrispettivo"))
    p.success_fee_perc = n(form.get("success_fee_perc"))
    p.fatturato = n(form.get("fatturato"))
    p.incassato = n(form.get("incassato"))
    conto = s(form.get("conto_incasso_id"))
    p.conto_incasso_id = int(conto) if conto else None
    p.note = s(form.get("note"))


@crm.get("/pratiche/nuova")
def nuova_pratica_form():
    p = Pratica(fase="Analisi fattibilità", priorita="Media")
    cliente = request.args.get("cliente", "")
    if cliente:
        p.cliente_id = int(cliente)
    bando_id = request.args.get("bando", "")
    if bando_id:
        bando = SessionLocale.get(Bando, int(bando_id))
        if bando:
            p.bando_id = bando.id
            p.nome_bando = bando.nome
            p.ente = bando.ente
            p.tipologia = bando.tipologia
            p.perc_contributo = bando.perc_contributo
            p.importo_max = bando.importo_max_testo or (
                f"{bando.contributo_max:.0f} €" if bando.contributo_max else None)
            p.data_apertura = bando.data_apertura
            p.data_scadenza = bando.data_scadenza
    return rendi("pratica_form", titolo="Nuova pratica", pagina="pratiche", p=p,
                 clienti=SessionLocale.query(Cliente).order_by(Cliente.ragione_sociale).all(),
                 bandi=SessionLocale.query(Bando).order_by(Bando.nome).all(),
                 conti=SessionLocale.query(ContoBancario).filter(
                     ContoBancario.attivo.is_(True)).order_by(ContoBancario.nome).all())


@crm.post("/pratiche/nuova")
def nuova_pratica():
    if not request.form.get("cliente_id") or not s(request.form.get("nome_bando")):
        avvisa("Servono il cliente e il nome del bando.", "ko")
        return redirect("/crm/pratiche/nuova")
    p = Pratica(codice=prossimo_codice(SessionLocale, Pratica, "PR"))
    _leggi_pratica(request.form, p)
    SessionLocale.add(p)
    SessionLocale.commit()
    avvisa(f"Pratica {p.codice} creata.")
    return redirect(f"/crm/pratiche/{p.id}")


@crm.get("/pratiche/<int:pid>")
def scheda_pratica(pid):
    p = SessionLocale.get(Pratica, pid)
    if not p:
        abort(404)
    if not p.token_caricamento:
        p.token_caricamento = secrets.token_urlsafe(24)
        SessionLocale.commit()
    return rendi("pratica", titolo=p.nome_bando, pagina="pratiche", p=p)


@crm.post("/pratiche/<int:pid>/invia-link")
def invia_link_pratica(pid):
    p = SessionLocale.get(Pratica, pid)
    destinatario = s(request.form.get("destinatario"))
    if not p or not destinatario:
        avvisa("Serve un indirizzo email valido.", "ko")
        return redirect(f"/crm/pratiche/{pid}")
    io_stesso = utente_corrente()
    link = f"{request.url_root.rstrip('/')}/crm/richiesta/{p.token_caricamento}"
    corpo = (
        f"Buongiorno,\n\n"
        f"per la pratica \"{p.nome_bando}\" ci servono alcuni documenti. Può caricarli da questo link, "
        f"anche in più volte:\n\n{link}\n\n"
        f"Grazie,\n{io_stesso.nome}\nEnergelia S.r.l."
    )
    ok, errore = invia_email(destinatario, f"Documenti per {p.nome_bando}", corpo,
                             rispondi_a=io_stesso.email, nome_mittente="Energelia")
    avvisa("Email inviata." if ok else f"Invio non riuscito: {errore}", "ok" if ok else "ko")
    return redirect(f"/crm/pratiche/{pid}")


@crm.post("/pratiche/<int:pid>/voci/nuova")
def nuova_voce_richiesta(pid):
    etichetta = s(request.form.get("etichetta"))
    if not etichetta:
        avvisa("La voce ha bisogno di un'etichetta.", "ko")
        return redirect(f"/crm/pratiche/{pid}")
    tipo = request.form.get("tipo_risposta") if request.form.get("tipo_risposta") in ("file", "testo") else "file"
    ordine = SessionLocale.query(func.count(VoceRichiesta.id)).filter_by(pratica_id=pid).scalar()
    SessionLocale.add(VoceRichiesta(pratica_id=pid, etichetta=etichetta, tipo_risposta=tipo, ordine=ordine))
    SessionLocale.commit()
    avvisa("Voce aggiunta.")
    return redirect(f"/crm/pratiche/{pid}")


@crm.post("/voci/<int:vid>/elimina")
def elimina_voce_richiesta(vid):
    v = SessionLocale.get(VoceRichiesta, vid)
    if v:
        pid = v.pratica_id
        SessionLocale.delete(v)
        SessionLocale.commit()
        return redirect(f"/crm/pratiche/{pid}")
    return redirect("/crm/pratiche")


# ------------------------------------------------------------------ bandi

def _leggi_bando(form, b):
    b.nome = s(form.get("nome"))
    b.ente = s(form.get("ente"))
    b.tipologia = s(form.get("tipologia"))
    b.dotazione = n(form.get("dotazione"))
    b.perc_contributo = n(form.get("perc_contributo"))
    b.contributo_max = n(form.get("contributo_max"))
    b.importo_max_testo = s(form.get("importo_max_testo"))
    b.data_apertura = d(form.get("data_apertura"))
    b.data_scadenza = d(form.get("data_scadenza"))
    b.chi_puo_partecipare = s(form.get("chi_puo_partecipare"))
    b.cosa_finanziabile = s(form.get("cosa_finanziabile"))
    b.spese_non_ammissibili = s(form.get("spese_non_ammissibili"))
    b.criteri = s(form.get("criteri"))
    b.fasi_tempi = s(form.get("fasi_tempi"))
    b.come_presentare = s(form.get("come_presentare"))
    b.perche_interessante = s(form.get("perche_interessante"))
    b.criticita = s(form.get("criticita"))


@crm.get("/bandi")
def lista_bandi():
    q = request.args.get("q", "")
    query = SessionLocale.query(Bando)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Bando.nome.ilike(like), Bando.ente.ilike(like)))
    elenco = query.order_by(Bando.nome).all()
    return rendi("bandi", titolo="Bandi", pagina="bandi", elenco=elenco, q=q,
                 configurato=anthropic_configurato())


@crm.get("/bandi/nuovo")
def nuovo_bando_form():
    return rendi("bando_form", titolo="Nuovo bando", pagina="bandi", b=Bando(),
                 configurato=anthropic_configurato(), tipologie=TIPOLOGIE)


@crm.post("/bandi/nuovo")
def nuovo_bando():
    if not s(request.form.get("nome")):
        avvisa("Serve almeno il nome del bando.", "ko")
        return redirect("/crm/bandi/nuovo")
    b = Bando()
    _leggi_bando(request.form, b)
    SessionLocale.add(b)
    SessionLocale.commit()
    avvisa(f"Bando \"{b.nome}\" salvato.")
    return redirect(f"/crm/bandi/{b.id}")


@crm.get("/bandi/<int:bid>")
def scheda_bando(bid):
    b = SessionLocale.get(Bando, bid)
    if not b:
        abort(404)
    return rendi("bando", titolo=b.nome, pagina="bandi", b=b, configurato=anthropic_configurato())


@crm.get("/bandi/<int:bid>/modifica")
def modifica_bando_form(bid):
    b = SessionLocale.get(Bando, bid)
    if not b:
        abort(404)
    return rendi("bando_form", titolo="Modifica bando", pagina="bandi", b=b,
                 configurato=anthropic_configurato(), tipologie=TIPOLOGIE)


@crm.post("/bandi/<int:bid>/modifica")
def modifica_bando(bid):
    b = SessionLocale.get(Bando, bid)
    if not b:
        abort(404)
    _leggi_bando(request.form, b)
    SessionLocale.commit()
    avvisa("Modifiche salvate.")
    return redirect(f"/crm/bandi/{b.id}")


@crm.post("/bandi/nuovo-da-pdf")
def nuovo_bando_da_pdf():
    if not anthropic_configurato():
        avvisa("L'estrazione automatica non è configurata: manca ANTHROPIC_API_KEY.", "ko")
        return redirect("/crm/bandi/nuovo")
    caricato = request.files.get("file")
    if not caricato or not caricato.filename:
        avvisa("Scegli un PDF da caricare.", "ko")
        return redirect("/crm/bandi/nuovo")
    try:
        testo = _estrai_testo_pdf(caricato.read())
        dati = estrai_bando_da_testo(testo)
    except Exception as errore:
        avvisa(f"Estrazione non riuscita: {errore}", "ko")
        return redirect("/crm/bandi/nuovo")

    b = Bando(testo_originale=testo)
    for campo in ("nome", "ente", "tipologia", "importo_max_testo", "chi_puo_partecipare",
                  "cosa_finanziabile", "spese_non_ammissibili", "criteri", "fasi_tempi",
                  "come_presentare", "perche_interessante", "criticita"):
        valore = (dati.get(campo) or "").strip()
        if valore:
            setattr(b, campo, valore)
    for campo in ("dotazione", "perc_contributo", "contributo_max"):
        valore = n(dati.get(campo))
        if valore is not None:
            setattr(b, campo, valore)
    for campo in ("data_apertura", "data_scadenza"):
        valore = d(dati.get(campo))
        if valore:
            setattr(b, campo, valore)
    if not b.nome:
        b.nome = caricato.filename.rsplit(".", 1)[0]

    SessionLocale.add(b)
    SessionLocale.commit()
    avvisa(f"Bando \"{b.nome}\" estratto dalla scheda. Controlla i campi prima di usarlo.")
    return redirect(f"/crm/bandi/{b.id}/modifica")


@crm.post("/bandi/<int:bid>/guida")
def genera_guida(bid):
    b = SessionLocale.get(Bando, bid)
    if not b:
        abort(404)
    if not anthropic_configurato():
        avvisa("La generazione guide non è configurata: manca ANTHROPIC_API_KEY.", "ko")
        return redirect(f"/crm/bandi/{bid}")
    try:
        b.guida_compilazione = genera_guida_compilazione(b)
        SessionLocale.commit()
        avvisa("Guida generata.")
    except Exception as errore:
        avvisa(f"Generazione non riuscita: {errore}", "ko")
    return redirect(f"/crm/bandi/{bid}")


@crm.post("/bandi/<int:bid>/approfondisci")
def approfondisci_bando(bid):
    b = SessionLocale.get(Bando, bid)
    if not b:
        abort(404)
    if not anthropic_configurato():
        avvisa("L'approfondimento non è configurato: manca ANTHROPIC_API_KEY.", "ko")
        return redirect(f"/crm/bandi/{bid}")
    try:
        b.approfondimento = genera_approfondimento(b)
        SessionLocale.commit()
        avvisa("Approfondimento generato.")
    except Exception as errore:
        avvisa(f"Generazione non riuscita: {errore}", "ko")
    return redirect(f"/crm/bandi/{bid}")


@crm.get("/bandi/<int:bid>/guida.txt")
def scarica_guida(bid):
    b = SessionLocale.get(Bando, bid)
    if not b or not b.guida_compilazione:
        abort(404)
    buffer = io.BytesIO(b.guida_compilazione.encode("utf-8"))
    nome_file = f"Guida_{b.nome}".replace(" ", "_").replace("/", "-") + ".txt"
    return send_file(buffer, as_attachment=True, download_name=nome_file, mimetype="text/plain")


@crm.get("/pratiche/<int:pid>/modifica")
def modifica_pratica_form(pid):
    p = SessionLocale.get(Pratica, pid)
    if not p:
        abort(404)
    return rendi("pratica_form", titolo="Modifica pratica", pagina="pratiche", p=p,
                 clienti=SessionLocale.query(Cliente).order_by(Cliente.ragione_sociale).all(),
                 conti=SessionLocale.query(ContoBancario).filter(
                     ContoBancario.attivo.is_(True)).order_by(ContoBancario.nome).all())


@crm.post("/pratiche/<int:pid>/modifica")
def modifica_pratica(pid):
    p = SessionLocale.get(Pratica, pid)
    if not p:
        abort(404)
    _leggi_pratica(request.form, p)
    SessionLocale.commit()
    avvisa("Modifiche salvate.")
    return redirect(f"/crm/pratiche/{p.id}")


# ----------------------------------------------------------------- attività

@crm.get("/attivita")
def lista_attivita():
    q = request.args.get("q", "")
    utente = request.args.get("utente", "")
    query = SessionLocale.query(Attivita).join(Cliente)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Attivita.testo.ilike(like), Cliente.ragione_sociale.ilike(like)))
    if utente:
        query = query.filter(Attivita.utente_id == int(utente))
    return rendi("attivita", titolo="Attività", pagina="attivita",
                 elenco=query.order_by(Attivita.data.desc(), Attivita.id.desc()).limit(200).all(),
                 q=q, utente_id=utente,
                 utenti=SessionLocale.query(Utente).order_by(Utente.nome).all())


@crm.post("/attivita/nuova")
def nuova_attivita():
    testo = s(request.form.get("testo"))
    cliente_id = int(request.form.get("cliente_id"))
    if not testo:
        avvisa("L'attività ha bisogno di un testo.", "ko")
        return redirect(f"/crm/clienti/{cliente_id}")
    pid = s(request.form.get("pratica_id"))
    a = Attivita(cliente_id=cliente_id,
                 pratica_id=int(pid) if pid else None,
                 data=d(request.form.get("data")) or dt.date.today(),
                 tipo=s(request.form.get("tipo")) or "Nota",
                 testo=testo,
                 utente_id=utente_corrente().id)
    SessionLocale.add(a)
    SessionLocale.commit()
    avvisa("Attività registrata.")
    return redirect(f"/crm/pratiche/{a.pratica_id}" if a.pratica_id else f"/crm/clienti/{a.cliente_id}")


# ------------------------------------------------------------- documenti

@crm.get("/documenti")
def lista_documenti():
    solo_da_smistare = request.args.get("tutti", "") != "1"
    query = SessionLocale.query(Documento)
    if solo_da_smistare:
        query = query.filter(Documento.stato == "da_smistare")
    elenco = query.order_by(Documento.creato_il.desc()).limit(200).all()
    return rendi("documenti", titolo="Documenti", pagina="documenti",
                 elenco=elenco, solo_da_smistare=solo_da_smistare,
                 configurato=drive_configurato())


@crm.post("/documenti/<int:did>/assegna")
def assegna_documento(did):
    doc = SessionLocale.get(Documento, did)
    pid = s(request.form.get("pratica_id"))
    if doc and pid:
        doc.pratica_id = int(pid)
        doc.stato = "assegnato"
        SessionLocale.commit()
        avvisa(f"{doc.nome_file} assegnato.")
    return redirect(request.referrer or "/crm/documenti")


# --------------------------------------------------------------------- lead

@crm.get("/lead")
def lista_lead():
    q = request.args.get("q", "")
    stato = request.args.get("stato", "")
    fonte = request.args.get("fonte", "")
    provincia = request.args.get("provincia", "")
    try:
        pagina_n = max(1, int(request.args.get("pagina", 1)))
    except ValueError:
        pagina_n = 1
    PER_PAGINA = 50

    query = SessionLocale.query(Lead)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Lead.nome.ilike(like), Lead.email.ilike(like),
                                 Lead.telefono.ilike(like), Lead.cellulare.ilike(like)))
    if stato:
        query = query.filter(Lead.stato == stato)
    if fonte:
        query = query.filter(Lead.fonte == fonte)
    if provincia:
        query = query.filter(Lead.provincia == provincia)

    totale = query.count()
    pagine_tot = max(1, (totale + PER_PAGINA - 1) // PER_PAGINA)
    pagina_n = min(pagina_n, pagine_tot)
    elenco = (query.order_by(Lead.creato_il.desc())
              .offset((pagina_n - 1) * PER_PAGINA).limit(PER_PAGINA).all())

    fonti = [r[0] for r in SessionLocale.query(Lead.fonte).filter(Lead.fonte.isnot(None))
             .distinct().order_by(Lead.fonte).all()]
    province = [r[0] for r in SessionLocale.query(Lead.provincia).filter(Lead.provincia.isnot(None))
                .distinct().order_by(Lead.provincia).all()]

    parti_query = []
    if q: parti_query.append(f"q={q}")
    if stato: parti_query.append(f"stato={stato}")
    if fonte: parti_query.append(f"fonte={fonte}")
    if provincia: parti_query.append(f"provincia={provincia}")

    return rendi("lead", titolo="Lead", pagina="lead", elenco=elenco, totale=totale,
                 q=q, stato=stato, fonte=fonte, fonti=fonti, provincia=provincia, province=province,
                 stati_lead=STATI_LEAD,
                 pagina_n=pagina_n, pagine_tot=pagine_tot,
                 query_senza_pagina="&".join(parti_query))


@crm.post("/lead/<int:lid>/stato")
def stato_lead(lid):
    l = SessionLocale.get(Lead, lid)
    nuovo_stato = s(request.form.get("stato"))
    if l and nuovo_stato in STATI_LEAD:
        l.stato = nuovo_stato
        SessionLocale.commit()
    return redirect(request.referrer or "/crm/lead")


class _CellaFoglio:
    __slots__ = ("value", "column")
    def __init__(self, value, column):
        self.value = value
        self.column = column


class _FoglioCSV:
    """Fa sembrare un CSV un foglio openpyxl (stessa interfaccia minima:
    max_row, ws[riga], ws.cell(riga, colonna)), così le funzioni di import
    scritte per gli xlsx leggono un CSV senza saperlo — nessuna duplicazione."""
    def __init__(self, righe):
        self.righe = righe
        self.max_row = len(righe)

    def __getitem__(self, indice_riga):
        riga = self.righe[indice_riga - 1] if 1 <= indice_riga <= len(self.righe) else []
        return [_CellaFoglio(v, i + 1) for i, v in enumerate(riga)]

    def cell(self, riga, colonna):
        r = self.righe[riga - 1] if 1 <= riga <= len(self.righe) else []
        valore = r[colonna - 1] if 1 <= colonna <= len(r) else None
        return _CellaFoglio(valore, colonna)


def _leggi_csv(contenuto_bytes):
    """Prova utf-8 (con o senza BOM), poi latin-1. Riconosce da sé ; o , come separatore."""
    for codifica in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            testo = contenuto_bytes.decode(codifica)
            break
        except UnicodeDecodeError:
            continue
    else:
        testo = contenuto_bytes.decode("utf-8", errors="replace")
    try:
        delimitatore = csv.Sniffer().sniff(testo[:2000], delimiters=";,").delimiter
    except csv.Error:
        delimitatore = ";" if testo[:2000].count(";") >= testo[:2000].count(",") else ","
    righe = [riga for riga in csv.reader(io.StringIO(testo), delimiter=delimitatore)]
    return _FoglioCSV(righe)


INTESTAZIONI_LEAD = {
    "nome": "nome", "ragione sociale": "nome",
    "tipo (maps)": "tipo", "tipo": "tipo",
    "query ricerca": "query_ricerca", "query/settore": "query_ricerca",
    "indirizzo": "indirizzo",
    "comune": "comune",
    "provincia": "provincia",
    "cap": "cap",
    "tel. fisso": "telefono", "tel fisso": "telefono", "telefono": "telefono",
    "cellulare": "cellulare",
    "sito web": "sito", "sito": "sito", "pagina trovata": "sito",
    "email": "email", "altre email": "altra_email", "pec": "pec",
}


def _pulisci_indirizzo(valore):
    """Gli scraper Maps a volte anteponono un carattere-icona non stampabile
    all'indirizzo (zona Unicode privata E000–F8FF): lo tolgo insieme a spazi vuoti."""
    testo = str(valore).strip()
    testo = "".join(ch for ch in testo if not (0xE000 <= ord(ch) <= 0xF8FF))
    return testo.strip(" \n\t") or None


@crm.post("/lead/importa")
def importa_lead():
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    caricato = request.files.get("file")
    if not caricato or not caricato.filename:
        avvisa("Scegli un file .xlsx o .csv da caricare.", "ko")
        return redirect(request.referrer or "/crm/impostazioni")
    contenuto = caricato.read()
    try:
        if caricato.filename.lower().endswith(".csv"):
            ws = _leggi_csv(contenuto)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(contenuto), data_only=True)
            ws = wb[wb.sheetnames[0]]
    except Exception as errore:
        avvisa(f"Il file non è leggibile: {errore}", "ko")
        return redirect(request.referrer or "/crm/impostazioni")

    fonte = s(request.form.get("fonte")) or caricato.filename.rsplit(".", 1)[0]

    # Trovo la riga di intestazione vera cercando fra le prime righe quella
    # che matcha più etichette conosciute — gli scraper mettono titolo e
    # contatore prima delle colonne vere.
    riga_intestazione, colonne = None, {}
    for indice_riga in range(1, min(15, ws.max_row) + 1):
        trovate = {}
        for cella in ws[indice_riga]:
            etichetta = str(cella.value or "").strip().lower()
            if etichetta in INTESTAZIONI_LEAD:
                trovate[INTESTAZIONI_LEAD[etichetta]] = cella.column
        if len(trovate) > len(colonne):
            riga_intestazione, colonne = indice_riga, trovate
    if not riga_intestazione or "nome" not in colonne:
        avvisa("Non trovo una colonna 'Nome' nel file: controlla che sia un export dello scraper.", "ko")
        return redirect(request.referrer or "/crm/impostazioni")

    esistenti = {(n.lower(), ind or "") for n, ind in
                 SessionLocale.query(Lead.nome, Lead.indirizzo).all()}
    nuovi = 0
    visti_in_questo_giro = set()

    for numero_riga in range(riga_intestazione + 1, ws.max_row + 1):
        nome = ws.cell(numero_riga, colonne["nome"]).value
        if not nome or not str(nome).strip():
            continue
        nome = str(nome).strip()

        valori = {}
        for campo, colonna in colonne.items():
            if campo == "nome":
                continue
            v = ws.cell(numero_riga, colonna).value
            valori[campo] = str(v).strip() if v not in (None, "") else None
        if valori.get("indirizzo"):
            valori["indirizzo"] = _pulisci_indirizzo(valori["indirizzo"])

        # Riga separatore di sezione (es. "Aosta (167 contatti)"): ha il nome
        # ma nessun altro campo valorizzato — la salto.
        if not any(valori.values()):
            continue

        chiave = (nome.lower(), valori.get("indirizzo") or "")
        if chiave in esistenti or chiave in visti_in_questo_giro:
            continue
        visti_in_questo_giro.add(chiave)

        SessionLocale.add(Lead(nome=nome, fonte=fonte, stato="nuovo", **valori))
        nuovi += 1

    SessionLocale.commit()
    avvisa(f"Importati {nuovi} lead da \"{fonte}\".")
    return redirect(request.referrer or "/crm/impostazioni")


# ------------------------------------------------------------- impostazioni

@crm.get("/impostazioni")
def impostazioni():
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    return rendi("impostazioni", titolo="Impostazioni", pagina="impostazioni",
                 utenti=SessionLocale.query(Utente).order_by(Utente.nome).all(),
                 conti_tutti=SessionLocale.query(ContoBancario).order_by(ContoBancario.nome).all())


@crm.post("/conti/nuovo")
def nuovo_conto():
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    nome = s(request.form.get("nome"))
    if not nome:
        avvisa("Il conto ha bisogno di un nome.", "ko")
    else:
        SessionLocale.add(ContoBancario(nome=nome,
                                        iban=(s(request.form.get("iban")) or "").upper().replace(" ", "") or None,
                                        banca=s(request.form.get("banca"))))
        SessionLocale.commit()
        avvisa(f"Conto {nome} aggiunto.")
    return redirect("/crm/impostazioni")


@crm.post("/conti/<int:cid>/stato")
def stato_conto(cid):
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    conto = SessionLocale.get(ContoBancario, cid)
    if conto:
        conto.attivo = not conto.attivo
        SessionLocale.commit()
        avvisa(f"{conto.nome} è ora {'attivo' if conto.attivo else 'disattivato'}.")
    return redirect("/crm/impostazioni")


@crm.post("/utenti/nuovo")
def nuovo_utente():
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    email = (s(request.form.get("email")) or "").lower()
    password = request.form.get("password") or ""
    if len(password) < 8:
        avvisa("La password deve avere almeno 8 caratteri.", "ko")
    elif SessionLocale.query(Utente).filter(func.lower(Utente.email) == email).first():
        avvisa(f"Esiste già un utente con l'email {email}.", "ko")
    else:
        SessionLocale.add(Utente(nome=s(request.form.get("nome")), email=email,
                                 password_hash=hash_pw(password),
                                 ruolo=s(request.form.get("ruolo")) or "consulente"))
        SessionLocale.commit()
        avvisa(f"Utente {email} creato.")
    return redirect("/crm/impostazioni")


@crm.post("/utenti/<int:uid>/stato")
def stato_utente(uid):
    io_stesso = utente_corrente()
    if not io_stesso.is_admin:
        return redirect("/crm/")
    u = SessionLocale.get(Utente, uid)
    if u and u.id != io_stesso.id:
        u.attivo = not u.attivo
        SessionLocale.commit()
        avvisa(f"{u.nome} è ora {'attivo' if u.attivo else 'disattivato'}.")
    return redirect("/crm/impostazioni")


@crm.post("/password")
def cambia_password():
    u = SessionLocale.get(Utente, utente_corrente().id)
    nuova = request.form.get("nuova") or ""
    if not verifica_pw(request.form.get("attuale") or "", u.password_hash):
        avvisa("La password attuale non è corretta.", "ko")
    elif len(nuova) < 8:
        avvisa("La nuova password deve avere almeno 8 caratteri.", "ko")
    else:
        u.password_hash = hash_pw(nuova)
        SessionLocale.commit()
        avvisa("Password aggiornata.")
    return redirect("/crm/impostazioni")


# -------------------------------------------------------- import ed export

INTESTAZIONI_CLIENTI = {
    "ragione sociale": "ragione_sociale", "partita iva / cf": "piva", "settore ateco": "ateco",
    "dimensione impresa": "dimensione", "città sede legale": "citta", "provincia": "provincia",
    "regione": "regione", "referente aziendale": "referente", "ruolo referente": "ruolo_referente",
    "telefono": "telefono", "email": "email", "pec": "pec", "canale acquisizione": "canale",
    "data primo contatto": "data_primo_contatto", "prossima azione": "prossima_azione", "note": "note",
}
INTESTAZIONI_PRATICHE = {
    "nome bando": "nome_bando", "ente erogatore": "ente", "tipologia agevolazione": "tipologia",
    "% contributo": "perc_contributo", "importo massimo erogabile (€)": "importo_max",
    "data apertura bando": "data_apertura", "data scadenza presentazione": "data_scadenza",
    "fase attuale": "fase", "data presentazione domanda": "data_presentazione",
    "data esito": "data_esito", "importo richiesto (€)": "importo_richiesto",
    "importo concesso (€)": "importo_concesso", "scadenza rendicontazione": "scadenza_rendicontazione",
    "documenti mancanti": "documenti_mancanti", "prossimo step": "prossimo_step", "priorità": "priorita",
}


def _mappa(ws, dizionario, riga_intestazioni=4):
    colonne = {}
    for cella in ws[riga_intestazioni]:
        etichetta = str(cella.value or "").strip().lower()
        if etichetta in dizionario:
            colonne[dizionario[etichetta]] = cella.column
    return colonne


@crm.post("/importa")
def importa():
    if not utente_corrente().is_admin:
        return redirect("/crm/")
    caricato = request.files.get("file")
    if not caricato or not caricato.filename:
        avvisa("Scegli un file .xlsx o .csv da caricare.", "ko")
        return redirect(request.referrer or "/crm/impostazioni")
    e_csv = caricato.filename.lower().endswith(".csv")
    contenuto = caricato.read()
    try:
        if e_csv:
            ws_clienti = _leggi_csv(contenuto)
            fogli = {"Clienti": ws_clienti}   # un CSV ha un solo "foglio": niente Pratiche
        else:
            wb = openpyxl.load_workbook(io.BytesIO(contenuto), data_only=True)
            fogli = {nome: wb[nome] for nome in wb.sheetnames}
    except Exception as errore:
        avvisa(f"Il file non è leggibile: {errore}", "ko")
        return redirect(request.referrer or "/crm/impostazioni")

    nuovi_clienti = nuove_pratiche = 0
    per_codice = {}

    if "Clienti" in fogli:
        ws = fogli["Clienti"]
        col = _mappa(ws, INTESTAZIONI_CLIENTI)
        for riga in range(5, ws.max_row + 1):
            ragione = ws.cell(riga, col.get("ragione_sociale", 2)).value
            if not ragione or not str(ragione).strip():
                continue
            codice_xlsx = str(ws.cell(riga, 1).value or "").strip()
            esistente = SessionLocale.query(Cliente).filter(
                func.lower(Cliente.ragione_sociale) == str(ragione).strip().lower()).first()
            if esistente:
                per_codice[codice_xlsx] = esistente
                continue
            c = Cliente(codice=prossimo_codice(SessionLocale, Cliente, "CL"),
                        ragione_sociale=str(ragione).strip())
            for campo, colonna in col.items():
                if campo == "ragione_sociale":
                    continue
                valore = ws.cell(riga, colonna).value
                if valore in (None, ""):
                    continue
                if campo == "data_primo_contatto":
                    setattr(c, campo, valore.date() if isinstance(valore, dt.datetime) else d(valore))
                else:
                    setattr(c, campo, str(valore).strip()[:250])
            SessionLocale.add(c)
            SessionLocale.flush()
            per_codice[codice_xlsx] = c
            nuovi_clienti += 1

    if "Pratiche" in fogli:
        ws = fogli["Pratiche"]
        col = _mappa(ws, INTESTAZIONI_PRATICHE)
        for riga in range(5, ws.max_row + 1):
            codice_cliente = str(ws.cell(riga, 2).value or "").strip()
            bando = ws.cell(riga, col.get("nome_bando", 4)).value
            cliente = per_codice.get(codice_cliente)
            if not bando or not cliente:
                continue
            if SessionLocale.query(Pratica).filter(
                    Pratica.cliente_id == cliente.id,
                    func.lower(Pratica.nome_bando) == str(bando).strip().lower()).first():
                continue
            p = Pratica(codice=prossimo_codice(SessionLocale, Pratica, "PR"), cliente_id=cliente.id,
                        nome_bando=str(bando).strip(), fase="Analisi fattibilità", priorita="Media")
            for campo, colonna in col.items():
                if campo == "nome_bando":
                    continue
                valore = ws.cell(riga, colonna).value
                if valore in (None, ""):
                    continue
                if campo in ("data_apertura", "data_scadenza", "data_presentazione",
                             "data_esito", "scadenza_rendicontazione"):
                    setattr(p, campo, valore.date() if isinstance(valore, dt.datetime) else d(valore))
                elif campo in ("importo_richiesto", "importo_concesso"):
                    setattr(p, campo, n(valore))
                elif campo == "perc_contributo":
                    # nel foglio la percentuale è una frazione (0,75): la riporto in punti
                    numero = n(valore)
                    if numero is not None:
                        setattr(p, campo, numero * 100 if numero <= 1 else numero)
                else:
                    setattr(p, campo, str(valore).strip()[:250])
            SessionLocale.add(p)
            SessionLocale.flush()
            nuove_pratiche += 1

    SessionLocale.commit()
    avvisa(f"Importati {nuovi_clienti} clienti e {nuove_pratiche} pratiche.")
    return redirect(request.referrer or "/crm/impostazioni")


@crm.get("/esporta.xlsx")
def esporta_xlsx():
    wb = openpyxl.Workbook()
    intestazione = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sfondo = PatternFill("solid", fgColor="1F4E78")

    def scrivi(ws, colonne, righe, larghezze):
        ws.append(colonne)
        for cella in ws[1]:
            cella.font = intestazione
            cella.fill = sfondo
            cella.alignment = Alignment(vertical="center")
        for r in righe:
            ws.append(r)
        for i, larghezza in enumerate(larghezze, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larghezza
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Clienti"
    scrivi(ws, ["Codice", "Ragione sociale", "P.IVA/CF", "ATECO", "Dimensione", "Città", "Prov.",
                "Regione", "Referente", "Ruolo", "Telefono", "Email", "PEC", "Consulente",
                "Canale", "Primo contatto", "Prossima azione", "N. pratiche", "Note"],
           [[c.codice, c.ragione_sociale, c.piva, c.ateco, c.dimensione, c.citta, c.provincia,
             c.regione, c.referente, c.ruolo_referente, c.telefono, c.email, c.pec,
             c.consulente.nome if c.consulente else None, c.canale, c.data_primo_contatto,
             c.prossima_azione, len(c.pratiche), c.note]
            for c in SessionLocale.query(Cliente).order_by(Cliente.codice).all()],
           [10, 34, 16, 10, 12, 16, 7, 14, 28, 22, 14, 26, 26, 18, 14, 14, 26, 10, 60])

    ws = wb.create_sheet("Pratiche")
    scrivi(ws, ["Codice", "Cliente", "Bando", "Ente", "Tipologia", "% contributo", "Importo max",
                "Apertura", "Scadenza", "Fase", "Presentazione", "Esito", "Richiesto", "Concesso",
                "Rendicontazione", "Priorità", "Prossimo step", "Documenti mancanti",
                "Corrispettivo", "Success fee %", "Fatturato", "Incassato"],
           [[p.codice, p.cliente.ragione_sociale, p.nome_bando, p.ente, p.tipologia,
             float(p.perc_contributo) if p.perc_contributo else None, p.importo_max,
             p.data_apertura, p.data_scadenza, p.fase, p.data_presentazione, p.data_esito,
             float(p.importo_richiesto) if p.importo_richiesto else None,
             float(p.importo_concesso) if p.importo_concesso else None,
             p.scadenza_rendicontazione, p.priorita, p.prossimo_step, p.documenti_mancanti,
             float(p.corrispettivo) if p.corrispettivo else None,
             float(p.success_fee_perc) if p.success_fee_perc else None,
             float(p.fatturato) if p.fatturato else None,
             float(p.incassato) if p.incassato else None]
            for p in SessionLocale.query(Pratica).order_by(Pratica.codice).all()],
           [10, 30, 30, 18, 20, 12, 24, 12, 12, 24, 14, 12, 14, 14, 16, 10, 26, 30, 14, 13, 12, 12])

    ws = wb.create_sheet("Attività")
    scrivi(ws, ["Data", "Cliente", "Pratica", "Tipo", "Utente", "Testo"],
           [[a.data, a.cliente.ragione_sociale, a.pratica.nome_bando if a.pratica else None,
             a.tipo, a.utente.nome if a.utente else None, a.testo]
            for a in SessionLocale.query(Attivita).order_by(Attivita.data.desc()).all()],
           [12, 30, 30, 14, 20, 90])

    for foglio in wb.worksheets:
        for riga in foglio.iter_rows(min_row=2):
            for cella in riga:
                cella.font = Font(name="Arial", size=10)
                if isinstance(cella.value, dt.date):
                    cella.number_format = "DD/MM/YYYY"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"CRM_Energelia_{dt.date.today():%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@crm.get("/esporta.csv")
def esporta_csv():
    buffer = io.StringIO()
    scrittore = csv.writer(buffer, delimiter=";")
    scrittore.writerow(["Codice", "Ragione sociale", "P.IVA", "Città", "Provincia",
                        "Referente", "Telefono", "Email", "PEC", "Canale", "N. pratiche"])
    for c in SessionLocale.query(Cliente).order_by(Cliente.codice).all():
        scrittore.writerow([c.codice, c.ragione_sociale, c.piva, c.citta, c.provincia,
                            c.referente, c.telefono, c.email, c.pec, c.canale, len(c.pratiche)])
    return Response(buffer.getvalue().encode("utf-8-sig"), mimetype="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="clienti.csv"'})


# --------------------------------------------------------------------------
# INNESTO NELL'APP DEL SITO
# --------------------------------------------------------------------------

def _chiave_sessione():
    """Deve restare identica fra i worker di gunicorn, altrimenti i cookie
    firmati da un worker non valgono per l'altro e l'accesso salta."""
    chiave = os.environ.get("CRM_SECRET_KEY")
    if chiave:
        return chiave
    print("[crm] CRM_SECRET_KEY non impostata: chiave derivata da DATABASE_URL. "
          "Impostala su Render.")
    return hashlib.sha256(("energelia-crm|" + DB_URL).encode()).hexdigest()


def _allinea_colonne():
    """create_all() crea solo le tabelle che mancano, non aggiunge colonne
    a quelle già esistenti. Qui confrontiamo il modello con il database reale
    e aggiungiamo da soli le colonne mancanti, senza toccare i dati già presenti."""
    from sqlalchemy import inspect, text
    ispettore = inspect(engine)
    tipo_sql = {
        "VARCHAR": lambda c: f"VARCHAR({c.type.length})" if getattr(c.type, "length", None) else "VARCHAR",
        "TEXT": lambda c: "TEXT",
        "INTEGER": lambda c: "INTEGER",
        "NUMERIC": lambda c: f"NUMERIC({c.type.precision},{c.type.scale})",
        "DATE": lambda c: "DATE",
        "DATETIME": lambda c: "TIMESTAMP",
        "BOOLEAN": lambda c: "BOOLEAN",
    }
    with engine.begin() as conn:
        for tabella in Base.metadata.sorted_tables:
            if not ispettore.has_table(tabella.name):
                continue
            info_presenti = {c["name"]: c for c in ispettore.get_columns(tabella.name)}
            for colonna in tabella.columns:
                if colonna.name not in info_presenti:
                    nome_tipo = colonna.type.__class__.__name__.upper()
                    costruttore = tipo_sql.get(nome_tipo, lambda c: "TEXT")
                    conn.execute(text(
                        f'ALTER TABLE {tabella.name} ADD COLUMN {colonna.name} {costruttore(colonna)}'
                    ))
                    print(f"[crm] Aggiunta colonna {tabella.name}.{colonna.name}")
                    continue
                # Il modello dice Text ma il database ha ancora un VARCHAR con un
                # limite (creato in una versione precedente): lo allargo. Un
                # ALTER a TEXT su Postgres è economico, non riscrive le righe.
                if colonna.type.__class__.__name__.upper() == "TEXT":
                    tipo_attuale = str(info_presenti[colonna.name]["type"]).upper()
                    if tipo_attuale != "TEXT":
                        conn.execute(text(
                            f'ALTER TABLE {tabella.name} ALTER COLUMN {colonna.name} TYPE TEXT'
                        ))
                        print(f"[crm] Allargata a TEXT la colonna {tabella.name}.{colonna.name} (era {tipo_attuale})")


def prepara_database():
    Base.metadata.create_all(engine)
    _allinea_colonne()
    if SessionLocale.query(func.count(Utente.id)).scalar() == 0:
        SessionLocale.add(Utente(nome="Alberto Augusti", email=ADMIN_EMAIL.lower(),
                                 password_hash=hash_pw(ADMIN_PASSWORD), ruolo="admin"))
        SessionLocale.commit()
        print(f"[crm] Creato l'amministratore {ADMIN_EMAIL}")
        if ADMIN_PASSWORD == "energelia2026":
            print("[crm] La password è quella predefinita: cambiala da /crm/impostazioni.")
    SessionLocale.remove()


def init_crm(app):
    """Aggancia il CRM all'app Flask del sito."""
    if not app.secret_key:
        app.secret_key = _chiave_sessione()
    app.permanent_session_lifetime = dt.timedelta(hours=12)
    app.register_blueprint(crm)

    @app.teardown_appcontext
    def _chiudi_sessione(_exc=None):
        SessionLocale.remove()

    try:
        prepara_database()
    except Exception as errore:                     # il sito deve partire comunque
        print(f"[crm] Errore in fase di preparazione del database: {errore}")
    return app
